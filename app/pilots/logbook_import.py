"""Pilot logbook import — parsing, normalisation, mapping, and execution."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Any

import openpyxl  # pyright: ignore[reportMissingImports]
from flask_babel import gettext as _  # pyright: ignore[reportMissingImports]
from flask_babel import pgettext  # pyright: ignore[reportMissingImports]

# ── Constants ─────────────────────────────────────────────────────────────────

TARGET_FIELDS: list[str] = [
    "date",
    "aircraft_type",
    "aircraft_registration",
    "departure_place",
    "departure_time",
    "arrival_place",
    "arrival_time",
    "pic_name",
    "night_time",
    "instrument_time",
    "cross_country",
    "landings_day",
    "landings_night",
    "single_pilot_se",
    "single_pilot_me",
    "multi_pilot",
    "function_pic",
    "function_copilot",
    "function_dual",
    "function_instructor",
    "remarks",
    # virtual target — used for import validation only, not stored
    "total_flight_time_check",
]

# Normalised source column name → target field.
# Positional disambiguation suffixes (_2, _3 …) are applied before lookup,
# so "time" is the first TIME column (departure) and "time_2" is the second (arrival).
_ALIASES: dict[str, str] = {
    # EASA standard logbook column names (normalised)
    "date dd/mm/yy": "date",
    "date": "date",
    "aircraft type": "aircraft_type",
    "type": "aircraft_type",
    "aircraft registration number": "aircraft_registration",
    "registration": "aircraft_registration",
    "reg": "aircraft_registration",
    "from": "departure_place",
    "departure": "departure_place",
    "dep": "departure_place",
    "time": "departure_time",  # first TIME → departure
    "time_2": "arrival_time",  # second TIME → arrival
    "departure time": "departure_time",
    "arrival time": "arrival_time",
    "to": "arrival_place",
    "arrival": "arrival_place",
    "arr": "arrival_place",
    "pic name": "pic_name",
    "pic name (if not student pilot)": "pic_name",
    "captain": "pic_name",
    # Landings group — first DAY / NIGHT pair
    "day": "landings_day",
    "night": "landings_night",
    # Operational conditions group — second DAY / NIGHT pair (positional suffix)
    "day_2": "ignore",  # cross-country day — not a logbook field
    "night_2": "night_time",  # night flying time
    # Aircraft category
    "se": "single_pilot_se",
    "single engine": "single_pilot_se",
    "single pilot se": "single_pilot_se",
    "me": "single_pilot_me",
    "multi engine": "single_pilot_me",
    "single pilot me": "single_pilot_me",
    "multi pilot": "multi_pilot",
    "mp": "multi_pilot",
    # Pilot function
    "pic": "function_pic",
    "p1": "function_pic",
    "co-pic": "function_copilot",
    "co-pilot": "function_copilot",
    "copilot": "function_copilot",
    "p2": "function_copilot",
    "dual received": "function_dual",
    "dual": "function_dual",
    "student": "function_dual",
    "instructor": "function_instructor",
    "fi": "function_instructor",
    # Cross-country time
    "cross-country": "cross_country",
    "cross country": "cross_country",
    "x-country": "cross_country",
    "xc": "cross_country",
    # Total flight time — imported only to validate against computed sum
    "total flight time": "total_flight_time_check",
    # Catch-all
    "total": "ignore",
    "no. istr. appr.": "ignore",
    "ifr approaches": "ignore",
    "page": "ignore",
    "line": "ignore",
    "remarks": "remarks",
    "notes": "remarks",
    "comments": "remarks",
    "instrument time": "instrument_time",
    "ifr": "instrument_time",
    "instrument": "instrument_time",
    "night time": "night_time",
    # Group-prefixed column names (e.g. Belgian EASA logbook with a span header row)
    "departure & arrival from": "departure_place",
    "departure & arrival time": "departure_time",
    "departure & arrival time_2": "arrival_time",
    "departure & arrival to": "arrival_place",
    "aircraft category se": "single_pilot_se",
    "aircraft category me": "single_pilot_me",
    "operational conditions cross-country": "cross_country",
    "operational conditions day": "ignore",
    "operational conditions night": "night_time",
    "pilot function pic": "function_pic",
    "pilot function co-pic": "function_copilot",
    "pilot function dual received": "function_dual",
    "page subtotals total flight time": "ignore",
    "page subtotals total flight time_2": "ignore",
    "page subtotals pic": "ignore",
    "page subtotals dual": "ignore",
    "page subtotals night": "ignore",
    "page subtotals landings – day": "ignore",
    "page subtotals landings – night": "ignore",
    "page subtotals formated date": "ignore",
    "page subtotals formated type": "ignore",
    "page subtotals days since flight": "ignore",
    "page subtotals daytime duration": "ignore",
    "page subtotals non pic or dual": "ignore",
    "formated date": "ignore",
    "formated type": "ignore",
    "days since flight": "ignore",
    "daytime duration": "ignore",
    "non pic or dual": "ignore",
    # Landings group (group-prefixed)
    "landings day": "landings_day",
    "landings night": "landings_night",
    # Aircraft type (multiline cell name normalised to single space)
    "aircraft type name, model, variant": "aircraft_type",
}

# ── Data structures ───────────────────────────────────────────────────────────


@dataclass
class ParsedFile:
    """Result of parsing an uploaded logbook file."""

    norm_cols: list[str]  # normalised + disambiguated column keys
    raw_cols: list[str]  # original column labels (for display)
    header_row_index: int  # 0-based row index of the detected header
    data_rows: list[list[Any]]  # all rows after the header (including subtotals)
    fingerprint: str  # SHA-256 of norm_cols


@dataclass
class MappingProposal:
    """Proposed column mapping and how it was derived."""

    mapping: dict[str, str]  # norm_col_key → target_field or "ignore"
    match_type: str  # "exact", "fuzzy", "alias"
    fuzzy_score: float = 0.0  # 0.0–1.0, only meaningful for "fuzzy"
    matched_mapping_id: int | None = None


@dataclass
class ImportResult:
    """Summary returned after executing an import."""

    imported: int = 0
    subtotals: int = 0
    skipped: list[tuple[int, str]] = field(default_factory=list)  # (row_num, reason)
    # (row_num, reason) — rows that matched an entry already in the pilot's logbook
    duplicates: list[tuple[int, str]] = field(default_factory=list)
    # (row_num, source_col, target_field, repr(raw)) — non-empty cells that couldn't parse
    parse_warnings: list[tuple[int, str, str, str]] = field(default_factory=list)
    # (row_num, source_total, computed_total) — rows where total ≠ sum of components
    total_mismatch_warnings: list[tuple[int, float, float]] = field(
        default_factory=list
    )
    has_opening_balance: bool = False


# ── Normalisation ─────────────────────────────────────────────────────────────


def _norm(text: str) -> str:
    """Strip, collapse whitespace, lower-case."""
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _disambiguate(names: list[str]) -> list[str]:
    """Append _2, _3 … to duplicate names to make them unique."""
    seen: dict[str, int] = {}
    result: list[str] = []
    for n in names:
        if n in seen:
            seen[n] += 1
            result.append(f"{n}_{seen[n]}")
        else:
            seen[n] = 1
            result.append(n)
    return result


def _fingerprint(norm_cols: list[str]) -> str:
    payload = json.dumps(norm_cols, ensure_ascii=False, sort_keys=False)
    return hashlib.sha256(payload.encode()).hexdigest()


# ── Header detection ──────────────────────────────────────────────────────────


def _is_header_row(row: list[Any]) -> bool:
    """True if ≥ 50 % of non-empty cells are non-numeric strings and ≥ 4 non-empty."""
    non_empty = [c for c in row if c is not None and str(c).strip()]
    if len(non_empty) < 4:
        return False
    string_like = [
        c for c in non_empty if isinstance(c, str) and not _is_numeric_str(str(c))
    ]
    return len(string_like) / len(non_empty) >= 0.5


def _is_numeric_str(s: str) -> bool:
    try:
        float(s)
        return True
    except ValueError:
        return False


def _header_alias_score(row: list[Any]) -> int:
    """Count how many cells in *row* match a known alias."""
    return sum(1 for c in row if c is not None and _norm(str(c)) in _ALIASES)


def _find_header_row(rows: list[list[Any]], max_scan: int = 20) -> int | None:
    """Return 0-based index of the best header row within the first max_scan rows.

    Many logbook templates have a group-label row (e.g. "DEPARTURE & ARRIVAL",
    "LANDINGS") above the actual column-header row.  Both pass _is_header_row,
    but the actual header row has far more alias matches.  We score every
    candidate and return the one with the highest score; ties go to the earlier
    row.  If no alias matches are found we fall back to the first text-like row.
    """
    best_idx: int | None = None
    best_score = -1
    first_text_row: int | None = None

    for i, row in enumerate(rows[:max_scan]):
        if not _is_header_row(row):
            continue
        if first_text_row is None:
            first_text_row = i
        score = _header_alias_score(row)
        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx if (best_idx is not None and best_score > 0) else first_text_row


def _trim_trailing_empty_cols(
    all_rows: list[list[Any]], max_scan: int = 50
) -> list[list[Any]]:
    """Trim columns beyond the rightmost non-empty value in the first max_scan rows.

    Excel templates often declare thousands of formatted-but-empty columns.
    Without trimming, every empty cell becomes a separate mapping entry.
    """
    max_col = 0
    for row in all_rows[:max_scan]:
        for j in range(len(row) - 1, -1, -1):
            if row[j] is not None and str(row[j]).strip():
                max_col = max(max_col, j + 1)
                break
    if max_col == 0:
        return (
            all_rows  # pragma: no cover — blank file rejected by header detection first
        )
    return [row[:max_col] for row in all_rows]


# ── Group-header detection ────────────────────────────────────────────────────


def _merge_label_map(ws: Any) -> dict[tuple[int, int], str]:
    """Return {(0-based row, 0-based col): group_label} for every merged region.

    The label is the value of the top-left cell of each merged range.
    Every cell in the range gets the same label so downstream code can do a
    simple lookup without forward-fill.
    """
    result: dict[tuple[int, int], str] = {}
    for mc in ws.merged_cells.ranges:
        val = ws.cell(mc.min_row, mc.min_col).value
        if val is None or not str(val).strip():
            continue
        label = str(val).strip()
        for r in range(mc.min_row - 1, mc.max_row):
            for c in range(mc.min_col - 1, mc.max_col):
                result[(r, c)] = label
    return result


def _group_labels_from_map(
    merge_map: dict[tuple[int, int], str], row_idx: int, width: int
) -> list[str]:
    """Return per-column group labels for *row_idx* using exact merge metadata."""
    return [merge_map.get((row_idx, c), "") for c in range(width)]


def _group_labels_heuristic(row: list[Any], width: int) -> list[str] | None:
    """Fallback for CSV: forward-fill a sparse row to infer group labels.

    Requires at least two non-empty values with a gap between them (so a single
    spanning title cell does not fire).  Returns None if the pattern is absent.
    """
    padded: list[Any] = list(row[:width]) + [None] * max(0, width - len(row))

    prev_nonempty_idx: int | None = None
    has_span = False
    for i, val in enumerate(padded):
        if val is not None and str(val).strip():
            if prev_nonempty_idx is not None and i > prev_nonempty_idx + 1:
                has_span = True
                break
            prev_nonempty_idx = i
    if not has_span:
        return None

    result: list[str] = []
    current = ""
    for val in padded:
        if val is not None and str(val).strip():
            current = str(val).strip()
        result.append(current)
    return result


def _apply_group_labels(group_labels: list[str], raw_header: list[str]) -> list[str]:
    """Prepend each non-empty group label to the corresponding column header."""
    result = []
    for i, col in enumerate(raw_header):
        label = group_labels[i] if i < len(group_labels) else ""
        if label and col:
            result.append(f"{label} {col}")
        elif label:
            result.append(label)
        else:
            result.append(col)
    return result


# ── File parsing ──────────────────────────────────────────────────────────────


def parse_file(data: bytes, filename: str) -> ParsedFile:
    """Parse bytes from an uploaded file into a ParsedFile.

    Raises ValueError with a user-friendly message on format errors.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        return _parse_excel(data, filename)
    if ext == "csv":
        return _parse_csv(data, filename)
    raise ValueError(
        f"Unsupported file format: .{ext} — please upload a .csv or .xlsx file."
    )


# English-only Excel tab names used to identify the logbook sheet.
# Locale-specific equivalents live in the translation catalogue, keyed under
# the "excel tab name" pgettext context (see _preferred_sheet_names below).
_PREFERRED_SHEET_NAMES_EN: frozenset[str] = frozenset(
    {
        "logbook",
        "log",
        "flights",
        "flight log",
        "flights log",
        "journal",
    }
)


def _preferred_sheet_names() -> frozenset[str]:
    """Return preferred sheet names merged with their active-locale translations.

    Uses pgettext with a dedicated context rather than plain _() — these
    keywords are matched against user-supplied Excel tab names, not
    displayed anywhere in the UI, and several of them (logbook/flight log,
    log/journal) intentionally translate to the same word in French. A
    context keeps that expected collision scoped to this lookup, so it can
    never silently mask a real duplicate-translation issue if the same
    English word is ever used as an actual UI label elsewhere.

    The context string is repeated as a literal at each call site (not a
    shared constant) because pybabel's extractor is a static source scanner —
    it can only resolve literal string arguments, not a variable reference.
    """
    t_logbook = pgettext("excel tab name", "logbook")
    t_log = pgettext("excel tab name", "log")
    t_flights = pgettext("excel tab name", "flights")
    t_flight_log = pgettext("excel tab name", "flight log")
    t_flights_log = pgettext("excel tab name", "flights log")
    t_journal = pgettext("excel tab name", "journal")
    translated = frozenset(
        {t_logbook, t_log, t_flights, t_flight_log, t_flights_log, t_journal}
    )
    return _PREFERRED_SHEET_NAMES_EN | frozenset(s.strip().lower() for s in translated)


def _pick_best_excel_sheet(wb: Any) -> tuple[str, list[list[Any]]]:
    """Return (sheet_name, all_rows) for the most likely logbook sheet.

    Prefers sheets whose name matches a known logbook keyword (fast path, reads
    only the chosen sheet).  When no name matches, reads every sheet and scores
    by alias matches in the detected header, returning the highest-scoring sheet
    together with its rows.  Falls back to the first sheet when scoring is
    inconclusive.

    Rows are returned here because openpyxl read-only worksheets use a one-shot
    generator — calling iter_rows() twice on the same worksheet yields nothing on
    the second call.
    """
    preferred = _preferred_sheet_names()
    for ws in wb.worksheets:
        if ws.title.strip().lower() in preferred:
            return ws.title, [list(row) for row in ws.iter_rows(values_only=True)]

    best_name = wb.worksheets[0].title if wb.worksheets else ""
    best_score = -1
    best_rows: list[list[Any]] = []
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        trimmed = _trim_trailing_empty_cols(list(rows))
        hi = _find_header_row(trimmed)
        score = _header_alias_score(trimmed[hi]) if hi is not None else 0
        if score > best_score:
            best_score = score
            best_name = ws.title
            best_rows = rows
    return best_name, best_rows


def _parse_excel(data: bytes, filename: str) -> ParsedFile:
    try:
        wb_ro = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not open Excel file: {exc}") from exc
    sheet_name, all_rows = _pick_best_excel_sheet(wb_ro)
    wb_ro.close()

    # Second load (non-read-only) to access merged cell ranges for exact group labels.
    excel_merge_map: dict[tuple[int, int], str] | None = None
    try:
        wb_full = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        excel_merge_map = _merge_label_map(wb_full[sheet_name])
        wb_full.close()
    except Exception:  # noqa: S110, BLE001  # fall back to heuristic in _build_parsed_file
        pass

    return _build_parsed_file(all_rows, filename, excel_merge_map=excel_merge_map)


def _parse_csv(data: bytes, filename: str) -> ParsedFile:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        all_rows: list[list[Any]] = list(reader)
    except csv.Error as exc:
        raise ValueError(f"Could not parse CSV file: {exc}") from exc
    return _build_parsed_file(all_rows, filename)


def _build_parsed_file(
    all_rows: list[list[Any]],
    filename: str,
    excel_merge_map: dict[tuple[int, int], str] | None = None,
) -> ParsedFile:
    all_rows = _trim_trailing_empty_cols(all_rows)
    header_idx = _find_header_row(all_rows)
    if header_idx is None:
        raise ValueError(
            "Could not detect a header row in this file. "
            "Make sure the file contains column names."
        )

    raw_header = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]

    # Prepend group labels from the row above the header, if one exists
    if header_idx > 0:
        width = len(raw_header)
        if excel_merge_map is not None:
            # Excel: exact boundaries from merged cell metadata
            _gl = _group_labels_from_map(excel_merge_map, header_idx - 1, width)
            group_labels: list[str] | None = _gl if any(_gl) else None
        else:
            # CSV: heuristic forward-fill (merged cell info unavailable)
            group_labels = _group_labels_heuristic(all_rows[header_idx - 1], width)
        if group_labels is not None:
            raw_header = _apply_group_labels(group_labels, raw_header)

    norm_raw = [_norm(c) for c in raw_header]
    norm_cols = _disambiguate(norm_raw)
    data_rows = all_rows[header_idx + 1 :]

    return ParsedFile(
        norm_cols=norm_cols,
        raw_cols=raw_header,
        header_row_index=header_idx,
        data_rows=data_rows,
        fingerprint=_fingerprint(norm_cols),
    )


# ── Mapping proposal ──────────────────────────────────────────────────────────


def propose_mapping(
    parsed: ParsedFile,
    existing_mappings: list[Any],  # list[LogbookImportMapping]
) -> MappingProposal:
    """Return the best mapping proposal for *parsed*, checking saved mappings first."""
    # 1. Exact fingerprint match
    for m in existing_mappings:
        if m.source_fingerprint == parsed.fingerprint:
            return MappingProposal(
                mapping=json.loads(m.column_mapping),
                match_type="exact",
                matched_mapping_id=m.id,
            )

    # 2. Fuzzy match — best overlap among saved mappings
    best_score = 0.0
    best_m = None
    new_set = set(parsed.norm_cols)
    for m in existing_mappings:
        saved_cols: list[str] = json.loads(m.source_columns)
        saved_set = set(saved_cols)
        if not saved_set:
            continue
        overlap = len(new_set & saved_set)
        score = overlap / max(len(new_set), len(saved_set))
        if score > best_score:
            best_score = score
            best_m = m

    if best_m is not None and best_score >= 0.6:
        saved_map: dict[str, str] = json.loads(best_m.column_mapping)
        # Build a mapping for the new columns, falling back to alias for unmatched ones
        merged = _alias_mapping(parsed.norm_cols)
        for col in parsed.norm_cols:
            if col in saved_map:
                merged[col] = saved_map[col]
        return MappingProposal(
            mapping=merged,
            match_type="fuzzy",
            fuzzy_score=best_score,
            matched_mapping_id=best_m.id,
        )

    # 3. Alias-only auto-mapping
    return MappingProposal(
        mapping=_alias_mapping(parsed.norm_cols),
        match_type="alias",
    )


def _alias_mapping(norm_cols: list[str]) -> dict[str, str]:
    """Apply the built-in alias table; unknown columns default to 'ignore'."""
    result: dict[str, str] = {}
    for col in norm_cols:
        result[col] = _ALIASES.get(col, "ignore")
    return result


# ── Subtotal detection ────────────────────────────────────────────────────────


def _is_subtotal_row(row: list[Any], date_col_idx: int | None) -> bool:
    if date_col_idx is None:
        return False
    if date_col_idx >= len(row):
        return True
    val = row[date_col_idx]
    if isinstance(val, timedelta):
        return True
    if val is None or (isinstance(val, str) and not val.strip()):
        return True
    return bool(isinstance(val, str) and "total" in val.lower())


# ── Value parsing ─────────────────────────────────────────────────────────────


def parse_date_value(val: Any) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        # Excel serial date number — openpyxl returns datetime; guard anyway
        return None
    if not isinstance(val, str):
        return None
    s = val.strip()
    if not s:
        return None
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_time_value(val: Any) -> time | None:
    """Parse a time-of-day value (HH:MM or Python time)."""
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if not isinstance(val, str):
        return None
    s = val.strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if m:
        try:
            return time(int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def parse_duration_value(val: Any) -> float | None:
    """Parse a duration into decimal hours (e.g. time(0,42) → 0.7, '1:24' → 1.4)."""
    if isinstance(val, timedelta):
        hours = val.total_seconds() / 3600
        return round(hours, 1) if hours >= 0 else None
    if isinstance(val, time):
        return round(val.hour + val.minute / 60, 1)
    if isinstance(val, datetime):
        # Excel sometimes returns a dummy date + the time-of-day
        t = val.time()
        return round(t.hour + t.minute / 60, 1)
    if isinstance(val, (int, float)):
        if not math.isfinite(val) or val < 0:
            return None
        return round(float(val), 1)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        m = re.match(r"^(\d+):(\d{2})$", s)
        if m:
            return round(int(m.group(1)) + int(m.group(2)) / 60, 1)
        try:
            v = float(s)
            return round(v, 1) if math.isfinite(v) and v >= 0 else None
        except ValueError:
            return None
    return None


def parse_int_value(val: Any) -> int | None:
    if isinstance(val, int):
        return val if val >= 0 else None
    if isinstance(val, float):
        try:
            return int(val) if val >= 0 else None
        except (ValueError, OverflowError):
            return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        try:
            n = int(float(s))
        except (ValueError, OverflowError):
            return None
        return n if n >= 0 else None
    return None


def _is_nonempty(val: Any) -> bool:
    """Return True when val carries a real value (not None, not blank string)."""
    if val is None:
        return False
    if isinstance(val, str):
        return bool(val.strip())
    return True


# Map target field → (human-readable type name, parser function)
_FIELD_TYPE: dict[str, tuple[str, Callable[[Any], Any]]] = {
    "date": ("date", parse_date_value),
    "departure_time": ("time", parse_time_value),
    "arrival_time": ("time", parse_time_value),
    "night_time": ("duration", parse_duration_value),
    "instrument_time": ("duration", parse_duration_value),
    "cross_country": ("duration", parse_duration_value),
    "total_flight_time_check": ("duration", parse_duration_value),
    "single_pilot_se": ("duration", parse_duration_value),
    "single_pilot_me": ("duration", parse_duration_value),
    "multi_pilot": ("duration", parse_duration_value),
    "function_pic": ("duration", parse_duration_value),
    "function_copilot": ("duration", parse_duration_value),
    "function_dual": ("duration", parse_duration_value),
    "function_instructor": ("duration", parse_duration_value),
    "landings_day": ("integer", parse_int_value),
    "landings_night": ("integer", parse_int_value),
}

_HINT_SAMPLE_ROWS = 5


def type_hints(parsed: ParsedFile, mapping: dict[str, str]) -> dict[str, str]:
    """Return {col: hint_text} for columns where sample data doesn't match the proposed type.

    Samples up to _HINT_SAMPLE_ROWS rows. Returns a hint when non-empty values are
    present but any of them fail to parse — indicating a likely mapping mismatch.
    """
    col_index = {col: i for i, col in enumerate(parsed.norm_cols)}
    hints: dict[str, str] = {}
    for col, target in mapping.items():
        if target not in _FIELD_TYPE:
            continue
        type_name, parser = _FIELD_TYPE[target]
        idx = col_index.get(col)
        if idx is None:
            continue
        sample = [
            row[idx]
            for row in parsed.data_rows[:_HINT_SAMPLE_ROWS]
            if idx < len(row) and _is_nonempty(row[idx])
        ]
        if not sample:
            continue
        failed = [v for v in sample if parser(v) is None]
        if failed:
            example = str(failed[0])[:30]
            hints[col] = (
                f"Sample data doesn't look like a {type_name} (e.g. {example!r})"
            )
    return hints


# ── Preview rows ──────────────────────────────────────────────────────────────


def preview_rows(
    parsed: ParsedFile,
    mapping: dict[str, str],
    n: int = 5,
) -> list[dict[str, Any]]:
    """Return up to *n* non-subtotal data rows mapped to target field names."""
    date_idx = _date_col_index(parsed.norm_cols, mapping)
    result: list[dict[str, Any]] = []
    for row in parsed.data_rows:
        if _is_subtotal_row(row, date_idx):
            continue
        mapped: dict[str, Any] = {}
        for i, col in enumerate(parsed.norm_cols):
            target = mapping.get(col, "ignore")
            if target == "ignore":
                continue
            mapped[target] = row[i] if i < len(row) else None
        result.append(mapped)
        if len(result) >= n:
            break
    return result


def _date_col_index(norm_cols: list[str], mapping: dict[str, str]) -> int | None:
    for i, col in enumerate(norm_cols):
        if mapping.get(col) == "date":
            return i
    return None


# ── Import execution ──────────────────────────────────────────────────────────


def _row_get(row: list[Any], col_index: dict[str, int], col: str) -> Any:
    i = col_index.get(col)
    return row[i] if i is not None and i < len(row) else None


def _parse_row_date(
    row: list[Any], mapping: dict[str, str], col_index: dict[str, int]
) -> date | None:
    for col, target in mapping.items():
        if target == "date":
            return parse_date_value(_row_get(row, col_index, col))
    return None


def _build_entry_kwargs(
    row: list[Any],
    mapping: dict[str, str],
    col_index: dict[str, int],
    date_val: date,
) -> tuple[dict[str, Any], float | None, list[tuple[str, str, str]]]:
    """Build Flight (standalone, aircraft_id NULL) field kwargs for one data
    row (identity/batch fields like pic_user_id are the caller's
    responsibility to add).

    Returns (kwargs, source_total_flight_time_check, parse_warnings), where
    parse_warnings is a list of (col, target, raw_repr) for non-empty cells
    that couldn't be parsed as their target field's type. Shared by the
    normal import pass and the near-match conflict finder below, so the two
    can never compute a row's fields differently.
    """
    kwargs: dict[str, Any] = {"date": date_val}
    source_total: float | None = None
    parse_warnings: list[tuple[str, str, str]] = []

    for col, target in mapping.items():
        if target in ("ignore", "date"):
            continue
        raw = _row_get(row, col_index, col)
        if target == "total_flight_time_check":
            if _is_nonempty(raw):
                source_total = parse_duration_value(raw)
            continue
        if target in _FIELD_TYPE:
            # NB: unpack into _type_name, not _ — callers of this module also
            # call gettext (imported as `_`) in the same function scope;
            # reassigning `_` would shadow it for the rest of that function
            # (Python's function-scoping, not block-scoping).
            _type_name, parser = _FIELD_TYPE[target]
            parsed_val = parser(raw)
            if parsed_val is None and _is_nonempty(raw):
                parse_warnings.append((col, target, repr(str(raw)[:40])))
            kwargs[target] = parsed_val
        elif target == "aircraft_type":
            val = str(raw).strip() if raw is not None else None
            kwargs["other_aircraft_type"] = val
            if val:
                from utils import (
                    resolve_aircraft_type_icao,  # pyright: ignore[reportMissingImports]
                )

                kwargs["other_aircraft_type_icao"] = resolve_aircraft_type_icao(val)
        elif target == "aircraft_registration":
            kwargs["other_aircraft_registration"] = (
                str(raw).strip() if raw is not None else None
            )
        elif target == "departure_place":
            kwargs["departure_icao"] = str(raw).strip() if raw is not None else None
        elif target == "arrival_place":
            kwargs["arrival_icao"] = str(raw).strip() if raw is not None else None
        elif target == "pic_name":
            kwargs["pic_name"] = str(raw).strip() if raw is not None else None
        elif target == "remarks":
            kwargs["notes"] = str(raw).strip() if raw is not None else None

    return kwargs, source_total, parse_warnings


def _num(v: Any) -> float | None:
    # Numeric columns come back as decimal.Decimal; the freshly parsed side
    # is always plain float (parse_duration_value). Decimal('0.7') != 0.7
    # (float) directly — comparing Decimals converted from float literals
    # hits binary-rounding mismatches — so normalise both sides to float
    # before building/comparing keys.
    return None if v is None else float(v)


def _dup_key(kwargs: dict[str, Any]) -> tuple[Any, ...]:
    """Exact-duplicate key: date + aircraft + duration + landings. Shared by
    execute_import (silently skips exact matches) and find_conflicting_rows
    (must not also flag them as a near-match conflict needing review)."""
    return (
        kwargs["date"],
        kwargs.get("other_aircraft_registration"),
        kwargs.get("single_pilot_se"),
        kwargs.get("single_pilot_me"),
        kwargs.get("multi_pilot"),
        kwargs.get("landings_day"),
        kwargs.get("landings_night"),
    )


def _fetch_existing_dedup_keys(pilot_user_id: int) -> set[tuple[Any, ...]]:
    from models import Aircraft, Flight, db  # pyright: ignore[reportMissingImports]
    from sqlalchemy import or_  # pyright: ignore[reportMissingImports]

    # Registration is Aircraft.registration for a managed-aircraft row, or
    # the free-text other_aircraft_registration for a standalone one.
    rows = (
        db.session.query(
            Flight.date,
            Flight.other_aircraft_registration,
            Aircraft.registration,
            Flight.single_pilot_se,
            Flight.single_pilot_me,
            Flight.multi_pilot,
            Flight.landings_day,
            Flight.landings_night,
        )
        .outerjoin(Aircraft, Aircraft.id == Flight.aircraft_id)
        .filter(
            or_(
                Flight.pic_user_id == pilot_user_id,
                Flight.second_crew_user_id == pilot_user_id,
            )
        )
    )
    return {
        (
            row[0],
            row[2] or row[1],
            _num(row[3]),
            _num(row[4]),
            _num(row[5]),
            row[6],
            row[7],
        )
        for row in rows
    }


def _assign_pilot_identity(kwargs: dict[str, Any], pilot_user_id: int) -> None:
    """Put *pilot_user_id* into the correct crew slot for one row (mutates
    *kwargs* in place), based on its function_* breakdown.

    A personal-logbook row's "PIC name" column (already in kwargs["pic_name"]
    by the time this runs) records the actual PIC's name as free text,
    independent of which slot the importing pilot occupies on that row. When
    the row shows dual-received or copilot time and no PIC time, the
    importing pilot was not the PIC — they belong in the second_crew_* slot
    (as student or copilot) instead, leaving pic_user_id unset so pic_name's
    free-text name remains the row's only PIC identity. Any other case (PIC
    time logged, instructor time logged, or no function breakdown at all —
    e.g. a source file with no "Pilot function" column) keeps pic_user_id as
    the importing pilot, matching this app's original single-slot behaviour.
    """
    from models import CrewRole  # pyright: ignore[reportMissingImports]

    function_pic = kwargs.get("function_pic")
    function_dual = kwargs.get("function_dual")
    function_copilot = kwargs.get("function_copilot")

    if not function_pic and function_dual:
        kwargs["second_crew_user_id"] = pilot_user_id
        kwargs["second_crew_role"] = CrewRole.STUDENT
    elif not function_pic and function_copilot:
        kwargs["second_crew_user_id"] = pilot_user_id
        kwargs["second_crew_role"] = CrewRole.COPILOT
    else:
        kwargs["pic_user_id"] = pilot_user_id


def execute_import(
    parsed: ParsedFile,
    mapping: dict[str, str],
    pilot_user_id: int,
    batch_id: int,
    opening_balance: dict[str, Any] | None = None,
    skip_row_nums: set[int] | None = None,
) -> ImportResult:
    """Create standalone Flight rows (aircraft_id NULL) from *parsed* using
    *mapping*.

    Returns an ImportResult describing what happened.  Entries are added to
    db.session but NOT committed — the caller commits after also saving the
    batch/mapping records. *skip_row_nums* (1-based, matching the row
    numbering used throughout this module) lets the caller carve out rows
    it's handling separately — e.g. near-match conflicts routed to the
    interactive review step in app/pilots/routes.py — so they're excluded
    entirely from this pass (not counted as imported, duplicate, or skipped).
    """
    from models import Flight, db  # pyright: ignore[reportMissingImports]

    result = ImportResult()
    date_idx = _date_col_index(parsed.norm_cols, mapping)
    col_index = {col: i for i, col in enumerate(parsed.norm_cols)}
    skip_row_nums = skip_row_nums or set()

    entries_to_add: list[Flight] = []

    # Duplicate detection: re-importing the same file — either by mistake, or
    # deliberately after appending new rows to the source spreadsheet, which
    # is the expected workflow for keeping an import-based logbook current —
    # must only add genuinely new rows, not double up everything already
    # imported. Match on date + aircraft + the duration/landings figures
    # rather than departure/arrival place & time: those are the fields that
    # actually drive currency and compliance totals, and they stay populated
    # even for entries whose route/time genuinely isn't captured (e.g. a
    # hand-entered or backfilled logbook row) — a key that depended on
    # route/time would silently fail to match those.
    existing_keys = _fetch_existing_dedup_keys(pilot_user_id)

    for row_num, row in enumerate(parsed.data_rows, start=1):
        if row_num in skip_row_nums:
            continue
        if _is_subtotal_row(row, date_idx):
            result.subtotals += 1
            continue

        date_val = _parse_row_date(row, mapping, col_index)
        if date_val is None:
            raw_date = (
                row[date_idx] if date_idx is not None and date_idx < len(row) else None
            )
            result.skipped.append((row_num, f"unparseable date: {raw_date!r}"))
            continue

        kwargs, source_total, parse_warnings = _build_entry_kwargs(
            row, mapping, col_index, date_val
        )
        _assign_pilot_identity(kwargs, pilot_user_id)
        kwargs["import_batch_id"] = batch_id
        kwargs["source"] = "import"
        for col, target, raw_repr in parse_warnings:
            result.parse_warnings.append((row_num, col, target, raw_repr))

        dup_key = _dup_key(kwargs)
        if dup_key in existing_keys:
            result.duplicates.append(
                (
                    row_num,
                    _(
                        "matches an entry already in your logbook "
                        "(same date, aircraft, duration and landings)"
                    ),
                )
            )
            continue
        existing_keys.add(dup_key)

        if source_total is not None:
            computed = round(
                sum(
                    float(kwargs.get(f) or 0)
                    for f in ("single_pilot_se", "single_pilot_me", "multi_pilot")
                ),
                1,
            )
            if abs(source_total - computed) >= 0.15:
                result.total_mismatch_warnings.append((row_num, source_total, computed))

        entries_to_add.append(Flight(**kwargs))

    result.imported = len(entries_to_add)
    for e in entries_to_add:
        db.session.add(e)

    # Opening balance — synthetic entry dated one day before earliest imported.
    # Only one opening-balance entry ever makes sense per pilot — re-importing
    # with opening-balance values filled in again must not create a second one.
    if opening_balance and any(v for v in opening_balance.values() if v):
        ob_exists = (
            db.session.query(Flight.id)
            .filter_by(pic_user_id=pilot_user_id, notes="Opening balance (imported)")
            .first()
            is not None
        )
        if ob_exists:
            result.duplicates.append(
                (
                    0,
                    _(
                        "opening balance: an opening balance entry already "
                        "exists for this pilot — not created again"
                    ),
                )
            )
        else:
            earliest = (
                min(e.date for e in entries_to_add) if entries_to_add else date.today()
            )
            from datetime import timedelta as _td

            balance_date = earliest - _td(days=1)
            balance_entry = Flight(
                pic_user_id=pilot_user_id,
                import_batch_id=batch_id,
                source="import",
                date=balance_date,
                notes="Opening balance (imported)",
                night_time=opening_balance.get("night_time"),
                instrument_time=opening_balance.get("instrument_time"),
                single_pilot_se=opening_balance.get("single_pilot_se"),
                single_pilot_me=opening_balance.get("single_pilot_me"),
                multi_pilot=opening_balance.get("multi_pilot"),
                function_pic=opening_balance.get("function_pic"),
                function_copilot=opening_balance.get("function_copilot"),
                function_dual=opening_balance.get("function_dual"),
                function_instructor=opening_balance.get("function_instructor"),
            )
            db.session.add(balance_entry)
            result.has_opening_balance = True

    return result


# ── Near-match conflict detection (possible corrections) ───────────────────────

_CANDIDATE_MIN_SCORE = 3
_CANDIDATE_DURATION_TOLERANCE = 1.0


@dataclass
class ConflictRow:
    """A parsed row that isn't an exact duplicate but scores highly enough
    against one or more existing entries to plausibly be an edited version
    of one of them — needs a human decision, not a guess."""

    row_num: int
    kwargs: dict[str, Any]
    candidates: list[tuple[float, int]]  # (score, existing_entry_id), best first


def _kwargs_duration(kwargs: dict[str, Any]) -> float | None:
    parts = [
        kwargs.get(f) for f in ("single_pilot_se", "single_pilot_me", "multi_pilot")
    ]
    vals = [float(p) for p in parts if p is not None]
    return round(sum(vals), 1) if vals else None


def _score_candidate(kwargs: dict[str, Any], existing: Any) -> float:
    """Score how likely *existing* (a Flight row) is the same real-world
    flight as *kwargs* (a freshly parsed row), across 7 points: registration,
    departure, arrival, departure time, arrival time, total duration,
    landings. A point only counts toward the score if both sides have data
    for it — missing data on either side is neutral, never a mismatch, so a
    row whose route/time isn't captured on one side can still be recognised
    via duration + landings alone. The two time points can be fractional —
    see services.time_band_matching.
    """
    from services.time_band_matching import (  # pyright: ignore[reportMissingImports]
        DEFAULT_OFFSET_HOURS,
        offset_ring_step_minutes,
        shift_time,
        time_band_score,
    )

    score: float = 0.0

    reg_new = (kwargs.get("other_aircraft_registration") or "").strip().upper()
    reg_old = (existing.display_registration or "").strip().upper()
    if reg_new and reg_old and reg_new == reg_old:
        score += 1

    dep_new = (kwargs.get("departure_icao") or "").strip().upper()
    dep_old = (existing.departure_icao or "").strip().upper()
    if dep_new and dep_old and dep_new == dep_old:
        score += 1

    arr_new = (kwargs.get("arrival_icao") or "").strip().upper()
    arr_old = (existing.arrival_icao or "").strip().upper()
    if arr_new and arr_old and arr_new == arr_old:
        score += 1

    # Same-pair (existing.departure_time/arrival_time set) is scored tight,
    # expecting a near-exact repeat of the same real instant, with a ring
    # width of 1/3 of the managed aircraft's flight_counter_offset — pilots
    # log to a fixed precision (0.1h/6min at the 0.3h default), so a flat
    # few-minutes tolerance would put an ordinary rounding difference in the
    # wrong ring. Falls back to the model's own 0.3h default when this row
    # has no managed aircraft to read a real offset from. Cross-pair
    # (existing only has takeoff_time/landing_time — an airframe-import
    # placeholder, e.g. a hand-entered row from before this pilot ever ran
    # a logbook import) is scored against that time shifted by this same
    # offset, the closest estimate available of the block-to-airborne time
    # gap — this needs a *real* managed-aircraft offset, so a standalone
    # existing row (no aircraft_id, e.g. a rental logged nowhere else) has
    # this comparison skipped entirely rather than guessing one.
    offset_hours = (
        float(existing.aircraft.flight_counter_offset) if existing.aircraft else None
    )
    same_pair_step = offset_ring_step_minutes(
        offset_hours if offset_hours is not None else DEFAULT_OFFSET_HOURS
    )

    dep_time_new = kwargs.get("departure_time")
    if existing.departure_time is not None:
        score += time_band_score(
            dep_time_new, (existing.departure_time,), same_pair_step
        )
    elif existing.takeoff_time is not None and offset_hours is not None:
        offset_minutes = round(offset_hours * 60)
        center = shift_time(existing.takeoff_time, -offset_minutes)
        score += time_band_score(
            dep_time_new, (center,), offset_ring_step_minutes(offset_hours)
        )

    arr_time_new = kwargs.get("arrival_time")
    if existing.arrival_time is not None:
        score += time_band_score(arr_time_new, (existing.arrival_time,), same_pair_step)
    elif existing.landing_time is not None and offset_hours is not None:
        offset_minutes = round(offset_hours * 60)
        center = shift_time(existing.landing_time, offset_minutes)
        score += time_band_score(
            arr_time_new, (center,), offset_ring_step_minutes(offset_hours)
        )

    dur_new = _kwargs_duration(kwargs)
    dur_old = existing.total_flight_time
    if (
        dur_new is not None
        and dur_old is not None
        and abs(dur_new - dur_old) <= _CANDIDATE_DURATION_TOLERANCE
    ):
        score += 1

    land_day = kwargs.get("landings_day")
    land_night = kwargs.get("landings_night")
    if (
        land_day is not None
        and land_night is not None
        and existing.landings_day is not None
        and existing.landings_night is not None
        and land_day == existing.landings_day
        and land_night == existing.landings_night
    ):
        score += 1

    return score


def find_conflicting_rows(
    parsed: ParsedFile,
    mapping: dict[str, str],
    pilot_user_id: int,
    exclude_row_nums: set[int] | None = None,
) -> list[ConflictRow]:
    """Find rows that aren't an exact duplicate but plausibly match an
    existing entry closely enough (score >= _CANDIDATE_MIN_SCORE) to need a
    human decision: keep the existing entry, overwrite it with the new data,
    or import as a genuinely separate new entry.

    Skips subtotal rows, rows with an unparseable date, and exact duplicates
    (same as execute_import's own dedup — an unmodified re-upload scores the
    maximum on every point, so without this check it would incorrectly be
    routed to review instead of being silently skipped), plus any row_num in
    *exclude_row_nums* — typically rows already resolved in an earlier pass
    of this same review.

    Candidates also include *unclaimed* rows (pic_user_id AND
    second_crew_user_id both NULL) on a managed aircraft belonging to one of
    the pilot's own tenants — typically a hand-entered airframe-log row from
    before this pilot ever ran a logbook import. Without this, those rows
    are invisible to both this check and execute_import's own dedup (both
    only look at rows already linked to *pilot_user_id*), so re-importing a
    personal logbook can never discover that the real-world flight is
    already logged from the airframe side — it silently creates a second,
    duplicate Flight row instead of surfacing a conflict to resolve.
    """
    from models import (  # pyright: ignore[reportMissingImports]
        Aircraft,
        Flight,
        TenantUser,
    )
    from sqlalchemy import and_, or_  # pyright: ignore[reportMissingImports]

    exclude_row_nums = exclude_row_nums or set()
    date_idx = _date_col_index(parsed.norm_cols, mapping)
    col_index = {col: i for i, col in enumerate(parsed.norm_cols)}
    existing_keys = _fetch_existing_dedup_keys(pilot_user_id)
    conflicts: list[ConflictRow] = []

    tenant_ids = [
        row.tenant_id for row in TenantUser.query.filter_by(user_id=pilot_user_id).all()
    ]

    for row_num, row in enumerate(parsed.data_rows, start=1):
        if row_num in exclude_row_nums:
            continue
        if _is_subtotal_row(row, date_idx):
            continue
        date_val = _parse_row_date(row, mapping, col_index)
        if date_val is None:
            continue

        kwargs, _source_total, _parse_warnings = _build_entry_kwargs(
            row, mapping, col_index, date_val
        )
        if _dup_key(kwargs) in existing_keys:
            continue  # exact duplicate — execute_import's own dedup handles this

        same_day_filters = [
            Flight.pic_user_id == pilot_user_id,
            Flight.second_crew_user_id == pilot_user_id,
        ]
        if tenant_ids:
            same_day_filters.append(
                and_(
                    Flight.pic_user_id.is_(None),
                    Flight.second_crew_user_id.is_(None),
                    Flight.aircraft_id.in_(
                        Aircraft.query.with_entities(Aircraft.id).filter(
                            Aircraft.tenant_id.in_(tenant_ids)
                        )
                    ),
                )
            )
        same_day = Flight.query.filter(
            or_(*same_day_filters),
            Flight.date == date_val,
        ).all()
        scored = [
            (score, existing.id)
            for existing in same_day
            if (score := _score_candidate(kwargs, existing)) >= _CANDIDATE_MIN_SCORE
        ]
        if scored:
            scored.sort(key=lambda t: -t[0])
            conflicts.append(
                ConflictRow(row_num=row_num, kwargs=kwargs, candidates=scored)
            )

    return conflicts


def link_entries_to_aircraft(entries: list[Any]) -> int:
    """Promote each standalone Flight row (aircraft_id NULL,
    other_aircraft_registration set) in *entries* to a managed aircraft, for
    any whose registration matches an Aircraft belonging to one of the row's
    pic/second-crew pilot's own tenants. Returns the count promoted. Caller
    must commit.

    Unified-model note: this used to create a brand-new FlightEntry +
    FlightCrew and link the pilot's existing PilotLogbookEntry to it via
    flight_id. There's only one row now, so promotion is just updating
    aircraft_id (+ clearing the free-text other_aircraft_* fields) on the
    row that already exists — no second row, no separate crew row (the
    row's pic_user_id/pic_name are already whatever the CSV import set).
    """
    from models import (  # pyright: ignore[reportMissingImports]
        Aircraft,
        TenantUser,
        User,
        db,
    )

    def _norm_reg(reg: str) -> str:
        return reg.upper().replace("-", "").replace(" ", "")

    ac_by_tenant: dict[int, dict[str, Any]] = {}
    for ac in Aircraft.query.all():
        ac_by_tenant.setdefault(ac.tenant_id, {})[_norm_reg(ac.registration)] = ac

    pilot_tenant_ids: dict[int, set[int]] = {}

    def _tenants_for_pilot(pilot_user_id: int) -> set[int]:
        if pilot_user_id not in pilot_tenant_ids:
            pilot_tenant_ids[pilot_user_id] = {
                row.tenant_id
                for row in TenantUser.query.filter_by(user_id=pilot_user_id).all()
            }
        return pilot_tenant_ids[pilot_user_id]

    def _place_icao(place: str | None) -> str:
        if not place:
            return "ZZZZ"
        clean = re.sub(r"[^A-Z0-9]", "", place.upper())[:4]
        return clean if len(clean) == 4 else "ZZZZ"

    promoted = 0
    for entry in entries:
        if entry.aircraft_id is not None or not entry.other_aircraft_registration:
            continue
        pilot_user_id = entry.pic_user_id or entry.second_crew_user_id
        if pilot_user_id is None:
            continue
        norm_reg = _norm_reg(entry.other_aircraft_registration)
        ac = None
        for tid in _tenants_for_pilot(pilot_user_id):
            ac = ac_by_tenant.get(tid, {}).get(norm_reg)
            if ac is not None:
                break
        if ac is None:
            continue

        entry.aircraft_id = ac.id
        entry.departure_icao = _place_icao(entry.departure_icao)
        entry.arrival_icao = _place_icao(entry.arrival_icao)
        entry.other_aircraft_type = None
        entry.other_aircraft_type_icao = None
        entry.other_aircraft_registration = None
        entry.source = "logbook_import"

        # Step 2 (docs/backlog.md "reconcile imports from either side"): the
        # true airframe-side fields (flight_time_counter_*/flight_time,
        # engine_time_counter_*/engine_time — the maintenance-relevant hour
        # figures) aren't known from a personal logbook and stay NULL, same
        # as everywhere else in this app "if a field is not filled, it's not
        # filled" — no supposition, interpolation, or guessed value gets
        # invented here. That's also what keeps the flight list's
        # needs-attention icon (source == "logbook_import" and flight_time
        # is None) lit until a human fills these in from a real airframe
        # source. (A previous version of this function seeded a rough
        # engine_time_counter estimate here, on the assumption that nothing
        # read it for maintenance tracking — services/component_limits.py
        # now sums engine_time_counter for engine/propeller TBO, so that
        # assumption no longer holds and the guess was removed.)

        if not entry.pic_name:
            pilot_user = db.session.get(User, pilot_user_id)
            if pilot_user:
                entry.pic_name = pilot_user.name or pilot_user.email

        promoted += 1

    return promoted
