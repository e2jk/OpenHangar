"""Tests for Phase 28: pilot logbook import (service layer + routes)."""

import io
import json
import os
import tempfile
from datetime import date, time, timedelta

import pw_hash as _pw_hash  # pyright: ignore[reportMissingImports]
import pytest

from models import (  # pyright: ignore[reportMissingImports]
    Aircraft,
    FlightEntry,
    LogbookImportBatch,
    LogbookImportMapping,
    PilotLogbookEntry,
    Role,
    Tenant,
    TenantUser,
    User,
    db,
)
from pilots.logbook_import import (  # pyright: ignore[reportMissingImports]
    ParsedFile,
    _alias_mapping,
    _apply_group_labels,
    _disambiguate,
    _fingerprint,
    _group_labels_from_map,
    _group_labels_heuristic,
    _is_subtotal_row,
    _merge_label_map,
    _norm,
    execute_import,
    parse_date_value,
    parse_duration_value,
    parse_file,
    parse_time_value,
    preview_rows,
    propose_mapping,
    type_hints,
)


# ── Fixtures ──────────────────────────────────────────────────────────────────


def _make_user(email="import@example.com"):
    user = User(
        email=email,
        password_hash=_pw_hash.hash("pw"),
        is_active=True,
    )
    db.session.add(user)
    db.session.flush()
    tenant = Tenant(name="Import Test Hangar")
    db.session.add(tenant)
    db.session.flush()
    db.session.add(TenantUser(user_id=user.id, tenant_id=tenant.id, role=Role.OWNER))
    db.session.commit()
    return user.id


def _login(client, uid):
    with client.session_transaction() as sess:
        sess["user_id"] = uid
        sess["pilot_access"] = True


def _make_parsed_file(cols: list[str], rows: list[list]) -> ParsedFile:
    return ParsedFile(
        norm_cols=cols,
        raw_cols=cols,
        header_row_index=0,
        data_rows=rows,
        fingerprint=_fingerprint(cols),
    )


# ── Service: normalisation & disambiguation ───────────────────────────────────


class TestNormalisation:
    def test_norm_strips_and_lowercases(self):
        assert _norm("  DATE dd/mm/yy  ") == "date dd/mm/yy"

    def test_norm_collapses_whitespace(self):
        assert _norm("AIRCRAFT  TYPE") == "aircraft type"

    def test_disambiguate_unique(self):
        assert _disambiguate(["a", "b", "c"]) == ["a", "b", "c"]

    def test_disambiguate_duplicates(self):
        result = _disambiguate(["time", "time", "night", "night"])
        assert result == ["time", "time_2", "night", "night_2"]

    def test_disambiguate_triple(self):
        result = _disambiguate(["x", "x", "x"])
        assert result == ["x", "x_2", "x_3"]

    def test_fingerprint_deterministic(self):
        cols = ["date", "from", "time", "time_2"]
        assert _fingerprint(cols) == _fingerprint(cols)

    def test_fingerprint_order_sensitive(self):
        assert _fingerprint(["a", "b"]) != _fingerprint(["b", "a"])


# ── Service: date / time / duration parsing ───────────────────────────────────


class TestParsers:
    def test_parse_date_datetime_object(self):
        from datetime import datetime

        assert parse_date_value(datetime(2024, 3, 15)) == date(2024, 3, 15)

    def test_parse_date_european(self):
        assert parse_date_value("15/03/24") == date(2024, 3, 15)
        assert parse_date_value("15/03/2024") == date(2024, 3, 15)

    def test_parse_date_iso(self):
        assert parse_date_value("2024-03-15") == date(2024, 3, 15)

    def test_parse_date_none_on_garbage(self):
        assert parse_date_value("not a date") is None
        assert parse_date_value(None) is None

    def test_parse_time_string(self):
        assert parse_time_value("09:30") == time(9, 30)
        assert parse_time_value("14:05") == time(14, 5)

    def test_parse_time_object(self):
        assert parse_time_value(time(8, 0)) == time(8, 0)

    def test_parse_time_none_on_garbage(self):
        assert parse_time_value("not:time") is None
        assert parse_time_value("") is None

    def test_parse_duration_time_object(self):
        assert parse_duration_value(time(0, 42)) == 0.7
        assert parse_duration_value(time(1, 24)) == 1.4

    def test_parse_duration_hhmm_string(self):
        assert parse_duration_value("1:24") == 1.4
        assert parse_duration_value("0:42") == 0.7

    def test_parse_duration_decimal_string(self):
        assert parse_duration_value("1.5") == 1.5

    def test_parse_duration_float(self):
        assert parse_duration_value(1.5) == 1.5

    def test_parse_duration_timedelta_as_hours(self):
        assert parse_duration_value(timedelta(hours=1, minutes=24)) == 1.4
        assert parse_duration_value(timedelta(minutes=42)) == 0.7
        assert parse_duration_value(timedelta(seconds=-1)) is None

    def test_parse_duration_none_on_empty(self):
        assert parse_duration_value("") is None
        assert parse_duration_value(None) is None


# ── Service: header detection & file parsing ──────────────────────────────────


class TestFileParsing:
    def _make_csv(self, rows: list[list]) -> bytes:
        lines = [",".join(str(c) for c in row) for row in rows]
        return "\n".join(lines).encode()

    def test_simple_csv_header_at_row0(self):
        data = self._make_csv(
            [
                ["Date", "From", "To", "Type", "Reg", "PIC"],
                ["15/03/24", "EBNM", "EBAW", "C172", "OO-TST", "Smith"],
            ]
        )
        pf = parse_file(data, "log.csv")
        assert pf.header_row_index == 0
        assert "date" in pf.norm_cols
        assert len(pf.data_rows) == 1

    def test_two_row_header_easa_style(self):
        # Row 0: group headers (non-column names, mostly merged → treated as non-header)
        # Row 1: real column names
        data = self._make_csv(
            [
                ["DEPARTURE & ARRIVAL", "", "", "LANDINGS", "", "AIRCRAFT CATEGORY"],
                ["DATE dd/mm/yy", "FROM", "TO", "DAY", "NIGHT", "SE"],
                ["01/01/24", "EBNM", "EBAW", "1", "0", "0.5"],
            ]
        )
        pf = parse_file(data, "log.csv")
        # Group header row has only a few non-empty cells → may or may not pass threshold
        # The actual column row should be found by row 1 at latest
        assert pf.header_row_index <= 1
        assert any("date" in c for c in pf.norm_cols)

    def test_duplicate_time_columns_disambiguated(self):
        data = self._make_csv(
            [
                ["Date", "From", "Time", "To", "Time", "SE"],
                ["01/01/24", "EBNM", "09:00", "EBAW", "10:30", "1.5"],
            ]
        )
        pf = parse_file(data, "log.csv")
        assert "time" in pf.norm_cols
        assert "time_2" in pf.norm_cols

    def test_no_header_raises(self):
        data = b"1.0,2.0,3.0\n4.0,5.0,6.0\n"
        with pytest.raises(ValueError, match="header"):
            parse_file(data, "log.csv")

    def test_unsupported_extension_raises(self):
        with pytest.raises(ValueError, match="Unsupported"):
            parse_file(b"data", "log.flightlog")


# ── Service: subtotal row detection ──────────────────────────────────────────


class TestSubtotalDetection:
    def test_timedelta_date_is_subtotal(self):
        row = [timedelta(hours=10), "EBNM", "EBAW"]
        assert _is_subtotal_row(row, date_col_idx=0)

    def test_none_date_is_subtotal(self):
        assert _is_subtotal_row([None, "x"], date_col_idx=0)

    def test_total_text_is_subtotal(self):
        assert _is_subtotal_row(["TOTAL", "x"], date_col_idx=0)
        assert _is_subtotal_row(["PAGE TOTAL", "x"], date_col_idx=0)

    def test_valid_date_not_subtotal(self):
        assert not _is_subtotal_row(["15/03/24", "x"], date_col_idx=0)

    def test_none_date_col_idx_never_subtotal(self):
        assert not _is_subtotal_row(["15/03/24", "x"], date_col_idx=None)


# ── Service: alias mapping ────────────────────────────────────────────────────


class TestAliasMapping:
    def test_from_maps_to_departure(self):
        m = _alias_mapping(["from"])
        assert m["from"] == "departure_place"

    def test_to_maps_to_arrival(self):
        m = _alias_mapping(["to"])
        assert m["to"] == "arrival_place"

    def test_time_maps_to_departure(self):
        m = _alias_mapping(["time"])
        assert m["time"] == "departure_time"

    def test_time_2_maps_to_arrival(self):
        m = _alias_mapping(["time_2"])
        assert m["time_2"] == "arrival_time"

    def test_se_maps_to_single_pilot_se(self):
        m = _alias_mapping(["se"])
        assert m["se"] == "single_pilot_se"

    def test_pic_maps_to_function_pic(self):
        m = _alias_mapping(["pic"])
        assert m["pic"] == "function_pic"

    def test_night_first_maps_to_landings_night(self):
        m = _alias_mapping(["day", "night"])
        assert m["day"] == "landings_day"
        assert m["night"] == "landings_night"

    def test_night_second_maps_to_night_time(self):
        m = _alias_mapping(["night", "night_2"])
        assert m["night_2"] == "night_time"

    def test_unknown_column_defaults_to_ignore(self):
        m = _alias_mapping(["no. istr. appr."])
        assert m["no. istr. appr."] == "ignore"

    def test_total_flight_time_maps_to_check(self):
        m = _alias_mapping(["total flight time"])
        assert m["total flight time"] == "total_flight_time_check"


# ── Service: mapping proposal (exact / fuzzy / alias) ────────────────────────


class TestMappingProposal:
    def _make_parsed(self, cols: list[str]) -> ParsedFile:
        return ParsedFile(
            norm_cols=cols,
            raw_cols=cols,
            header_row_index=0,
            data_rows=[],
            fingerprint=_fingerprint(cols),
        )

    def test_exact_match_uses_saved_mapping(self, app):
        cols = ["date", "from", "to", "se"]
        pf = self._make_parsed(cols)

        class FakeMapping:
            source_fingerprint = pf.fingerprint
            column_mapping = json.dumps(
                {
                    "date": "date",
                    "from": "departure_place",
                    "to": "arrival_place",
                    "se": "single_pilot_se",
                }
            )
            source_columns = json.dumps(cols)
            id = 99

        proposal = propose_mapping(pf, [FakeMapping()])
        assert proposal.match_type == "exact"
        assert proposal.mapping["from"] == "departure_place"

    def test_fuzzy_match_applied_above_threshold(self, app):
        saved_cols = ["date", "from", "to", "se", "me", "pic"]
        new_cols = ["date", "from", "to", "se", "me", "pic_name"]  # 5/6 overlap ≈ 83%

        pf = self._make_parsed(new_cols)

        class FakeMapping:
            source_fingerprint = "different_hash"
            column_mapping = json.dumps({c: "ignore" for c in saved_cols})
            source_columns = json.dumps(saved_cols)
            id = 1

        proposal = propose_mapping(pf, [FakeMapping()])
        assert proposal.match_type == "fuzzy"
        assert proposal.fuzzy_score >= 0.6

    def test_no_fuzzy_match_below_threshold(self, app):
        saved_cols = ["x", "y", "z"]
        new_cols = ["date", "from", "to", "se"]

        pf = self._make_parsed(new_cols)

        class FakeMapping:
            source_fingerprint = "different_hash"
            column_mapping = json.dumps({c: "ignore" for c in saved_cols})
            source_columns = json.dumps(saved_cols)
            id = 1

        proposal = propose_mapping(pf, [FakeMapping()])
        assert proposal.match_type == "alias"

    def test_no_saved_mappings_uses_alias(self, app):
        pf = self._make_parsed(["date", "from", "to"])
        proposal = propose_mapping(pf, [])
        assert proposal.match_type == "alias"
        assert proposal.mapping["from"] == "departure_place"


# ── Service: execute_import ───────────────────────────────────────────────────


class TestExecuteImport:
    def _make_parsed(self, rows, cols=None) -> ParsedFile:
        if cols is None:
            cols = ["date", "from", "to", "se", "pic"]
        return ParsedFile(
            norm_cols=cols,
            raw_cols=cols,
            header_row_index=0,
            data_rows=rows,
            fingerprint=_fingerprint(cols),
        )

    def test_basic_import_creates_entries(self, app):
        with app.app_context():
            uid = _make_user("exec1@example.com")
            # Create a placeholder batch id by inserting a real batch
            from datetime import datetime, timezone

            batch = LogbookImportBatch(
                pilot_user_id=uid,
                source_filename="test.csv",
                imported_at=datetime.now(timezone.utc),
            )
            db.session.add(batch)
            db.session.flush()
            bid = batch.id

            parsed = self._make_parsed(
                [
                    ["15/03/24", "EBNM", "EBAW", "0.5", "0.5"],
                    ["16/03/24", "EBAW", "EBNM", "0.8", "0.8"],
                ]
            )
            mapping = {
                "date": "date",
                "from": "departure_place",
                "to": "arrival_place",
                "se": "single_pilot_se",
                "pic": "function_pic",
            }
            result = execute_import(parsed, mapping, uid, bid)
            db.session.commit()

            assert result.imported == 2
            assert result.subtotals == 0
            assert result.skipped == []
            entries = PilotLogbookEntry.query.filter_by(
                pilot_user_id=uid, import_batch_id=bid
            ).all()
            assert len(entries) == 2
            assert entries[0].source == "import"

    def test_subtotal_rows_skipped(self, app):
        with app.app_context():
            uid = _make_user("exec2@example.com")
            from datetime import datetime, timezone

            batch = LogbookImportBatch(
                pilot_user_id=uid,
                source_filename="t.csv",
                imported_at=datetime.now(timezone.utc),
            )
            db.session.add(batch)
            db.session.flush()

            parsed = self._make_parsed(
                [
                    ["15/03/24", "EBNM", "EBAW", "0.5", "0.5"],
                    [timedelta(hours=10), "subtotal", "", "10", "10"],  # subtotal row
                ]
            )
            mapping = {
                "date": "date",
                "from": "departure_place",
                "to": "arrival_place",
                "se": "single_pilot_se",
                "pic": "function_pic",
            }
            result = execute_import(parsed, mapping, uid, batch.id)
            db.session.commit()

            assert result.imported == 1
            assert result.subtotals == 1

    def test_unparseable_date_skipped(self, app):
        with app.app_context():
            uid = _make_user("exec3@example.com")
            from datetime import datetime, timezone

            batch = LogbookImportBatch(
                pilot_user_id=uid,
                source_filename="t.csv",
                imported_at=datetime.now(timezone.utc),
            )
            db.session.add(batch)
            db.session.flush()

            parsed = self._make_parsed(
                [
                    ["not-a-date", "EBNM", "EBAW", "0.5", "0.5"],
                ]
            )
            mapping = {
                "date": "date",
                "from": "departure_place",
                "to": "arrival_place",
                "se": "single_pilot_se",
                "pic": "function_pic",
            }
            result = execute_import(parsed, mapping, uid, batch.id)
            db.session.commit()

            assert result.imported == 0
            assert len(result.skipped) == 1
            assert "unparseable date" in result.skipped[0][1]

    def test_opening_balance_creates_extra_entry(self, app):
        with app.app_context():
            uid = _make_user("exec4@example.com")
            from datetime import datetime, timezone

            batch = LogbookImportBatch(
                pilot_user_id=uid,
                source_filename="t.csv",
                imported_at=datetime.now(timezone.utc),
            )
            db.session.add(batch)
            db.session.flush()

            parsed = self._make_parsed(
                [
                    ["15/03/24", "EBNM", "EBAW", "0.5", "0.5"],
                ]
            )
            mapping = {
                "date": "date",
                "from": "departure_place",
                "to": "arrival_place",
                "se": "single_pilot_se",
                "pic": "function_pic",
            }
            ob = {"single_pilot_se": 100.0, "function_pic": 100.0}
            result = execute_import(parsed, mapping, uid, batch.id, opening_balance=ob)
            db.session.commit()

            assert result.has_opening_balance
            entries = PilotLogbookEntry.query.filter_by(
                pilot_user_id=uid, import_batch_id=batch.id
            ).all()
            # 1 real + 1 opening balance
            assert len(entries) == 2
            ob_entry = next(
                e for e in entries if e.remarks == "Opening balance (imported)"
            )
            assert ob_entry.date == date(2024, 3, 14)  # one day before 15/03/24
            assert float(ob_entry.single_pilot_se) == 100.0

    def test_duration_from_time_object(self, app):
        with app.app_context():
            uid = _make_user("exec5@example.com")
            from datetime import datetime, timezone

            batch = LogbookImportBatch(
                pilot_user_id=uid,
                source_filename="t.csv",
                imported_at=datetime.now(timezone.utc),
            )
            db.session.add(batch)
            db.session.flush()

            parsed = self._make_parsed(
                [
                    [date(2024, 3, 15), "EBNM", "EBAW", time(0, 42), time(0, 42)],
                ]
            )
            mapping = {
                "date": "date",
                "from": "departure_place",
                "to": "arrival_place",
                "se": "single_pilot_se",
                "pic": "function_pic",
            }
            execute_import(parsed, mapping, uid, batch.id)
            db.session.commit()

            entry = PilotLogbookEntry.query.filter_by(
                pilot_user_id=uid, import_batch_id=batch.id
            ).first()
            assert entry is not None
            assert float(entry.single_pilot_se) == 0.7

    def test_aircraft_type_icao_resolved_on_import(self, app):
        """Importing a row with a known aircraft_type populates aircraft_type_icao."""
        from datetime import date as _date, datetime, timezone

        with app.app_context():
            uid = _make_user("exec_icao@example.com")
            batch = LogbookImportBatch(
                pilot_user_id=uid,
                source_filename="icao.csv",
                imported_at=datetime.now(timezone.utc),
            )
            db.session.add(batch)
            db.session.flush()
            bid = batch.id

            parsed = self._make_parsed(
                [[_date(2024, 3, 15), "EBNM", "EBAW", "0.5", "0.5", "C172"]],
                cols=["date", "from", "to", "se", "pic", "type"],
            )
            mapping = {
                "date": "date",
                "from": "departure_place",
                "to": "arrival_place",
                "se": "single_pilot_se",
                "pic": "function_pic",
                "type": "aircraft_type",
            }
            execute_import(parsed, mapping, uid, bid)
            db.session.commit()

            entry = PilotLogbookEntry.query.filter_by(
                pilot_user_id=uid, import_batch_id=bid
            ).first()
            assert entry is not None
            assert entry.aircraft_type == "C172"
            assert entry.aircraft_type_icao == "C172"


# ── Service: parse warnings & type hints ─────────────────────────────────────


class TestParseWarnings:
    def _make_batch(self, uid: int) -> int:
        from datetime import datetime, timezone

        from models import LogbookImportBatch, db

        batch = LogbookImportBatch(
            pilot_user_id=uid,
            source_filename="x.csv",
            imported_at=datetime.now(timezone.utc),
            row_count=0,
            subtotal_count=0,
            skipped_count=0,
        )
        db.session.add(batch)
        db.session.flush()
        return batch.id

    def test_unparseable_duration_recorded(self, app):
        from datetime import date as _date

        with app.app_context():
            uid = _make_user("pw@example.com")
            bid = self._make_batch(uid)
            parsed = _make_parsed_file(
                ["date", "se"],
                [[_date(2024, 1, 1), "not a duration"]],
            )
            mapping = {"date": "date", "se": "single_pilot_se"}
            result = execute_import(parsed, mapping, uid, bid)
            assert len(result.parse_warnings) == 1
            row_num, col, target, raw = result.parse_warnings[0]
            assert row_num == 1
            assert target == "single_pilot_se"
            assert "not a duration" in raw

    def test_empty_cell_not_warned(self, app):
        from datetime import date as _date

        with app.app_context():
            uid = _make_user("pw2@example.com")
            bid = self._make_batch(uid)
            parsed = _make_parsed_file(
                ["date", "se"],
                [[_date(2024, 1, 1), None]],
            )
            mapping = {"date": "date", "se": "single_pilot_se"}
            result = execute_import(parsed, mapping, uid, bid)
            assert result.parse_warnings == []

    def test_cross_country_imported_when_present(self, app):
        from datetime import date as _date

        with app.app_context():
            uid = _make_user("xc1@example.com")
            bid = self._make_batch(uid)
            parsed = _make_parsed_file(
                ["date", "se", "cross-country"],
                [[_date(2024, 3, 15), "1.0", "0.8"]],
            )
            mapping = {
                "date": "date",
                "se": "single_pilot_se",
                "cross-country": "cross_country",
            }
            result = execute_import(parsed, mapping, uid, bid)
            db.session.commit()
            assert result.imported == 1
            entry = PilotLogbookEntry.query.filter_by(pilot_user_id=uid).first()
            assert entry is not None
            assert float(entry.cross_country) == 0.8

    def test_cross_country_null_when_absent(self, app):
        from datetime import date as _date

        with app.app_context():
            uid = _make_user("xc2@example.com")
            bid = self._make_batch(uid)
            parsed = _make_parsed_file(
                ["date", "se"],
                [[_date(2024, 3, 15), "1.0"]],
            )
            mapping = {"date": "date", "se": "single_pilot_se"}
            execute_import(parsed, mapping, uid, bid)
            db.session.commit()
            entry = PilotLogbookEntry.query.filter_by(pilot_user_id=uid).first()
            assert entry is not None
            assert entry.cross_country is None

    def test_total_mismatch_warning_when_sum_differs(self, app):
        from datetime import date as _date

        with app.app_context():
            uid = _make_user("tot1@example.com")
            bid = self._make_batch(uid)
            parsed = _make_parsed_file(
                ["date", "se", "total flight time"],
                [[_date(2024, 3, 15), "0.5", "1.5"]],  # sum=0.5, total=1.5
            )
            mapping = {
                "date": "date",
                "se": "single_pilot_se",
                "total flight time": "total_flight_time_check",
            }
            result = execute_import(parsed, mapping, uid, bid)
            assert len(result.total_mismatch_warnings) == 1
            row_num, src, comp = result.total_mismatch_warnings[0]
            assert row_num == 1
            assert src == 1.5
            assert comp == 0.5

    def test_no_total_mismatch_when_sum_matches(self, app):
        from datetime import date as _date

        with app.app_context():
            uid = _make_user("tot2@example.com")
            bid = self._make_batch(uid)
            parsed = _make_parsed_file(
                ["date", "se", "total flight time"],
                [[_date(2024, 3, 15), "1.5", "1.5"]],
            )
            mapping = {
                "date": "date",
                "se": "single_pilot_se",
                "total flight time": "total_flight_time_check",
            }
            result = execute_import(parsed, mapping, uid, bid)
            assert result.total_mismatch_warnings == []


class TestTypeHints:
    def test_bad_duration_column_flagged(self):
        parsed = _make_parsed_file(
            ["date", "se"],
            [["01/01/2024", "Excellent"], ["02/01/2024", "Good"]],
        )
        hints = type_hints(parsed, {"date": "date", "se": "single_pilot_se"})
        assert "se" in hints
        assert "duration" in hints["se"]

    def test_valid_duration_column_no_hint(self):
        parsed = _make_parsed_file(
            ["date", "se"],
            [["01/01/2024", "1:30"], ["02/01/2024", "0:45"]],
        )
        hints = type_hints(parsed, {"date": "date", "se": "single_pilot_se"})
        assert "se" not in hints

    def test_empty_column_no_hint(self):
        parsed = _make_parsed_file(
            ["date", "se"],
            [["01/01/2024", None], ["02/01/2024", ""]],
        )
        hints = type_hints(parsed, {"date": "date", "se": "single_pilot_se"})
        assert "se" not in hints

    def test_ignored_column_no_hint(self):
        parsed = _make_parsed_file(["date", "notes"], [["01/01/2024", "free text"]])
        hints = type_hints(parsed, {"date": "date", "notes": "remarks"})
        assert "notes" not in hints

    def test_mapped_col_not_in_norm_cols_ignored(self):
        # mapping references a column that isn't in the ParsedFile — should be skipped
        parsed = _make_parsed_file(["date"], [["01/01/2024"]])
        hints = type_hints(parsed, {"date": "date", "ghost_col": "single_pilot_se"})
        assert "ghost_col" not in hints

    def test_nonempty_non_str_non_none_triggers_warning(self, app):
        # A non-str, non-None value (e.g. int) that fails to parse should be warned.
        # This exercises the _is_nonempty `return True` branch for non-str/non-None.
        from datetime import date as _date

        with app.app_context():
            from datetime import datetime, timezone

            uid = _make_user("isnonempty@example.com")
            from models import LogbookImportBatch, db

            batch = LogbookImportBatch(
                pilot_user_id=uid,
                source_filename="x.csv",
                imported_at=datetime.now(timezone.utc),
                row_count=0,
                subtotal_count=0,
                skipped_count=0,
            )
            db.session.add(batch)
            db.session.flush()
            bid = batch.id

            # -999 is non-str, non-None, but parse_duration_value rejects negatives
            parsed = _make_parsed_file(
                ["date", "se"],
                [[_date(2024, 1, 1), -999]],
            )
            result = execute_import(
                parsed, {"date": "date", "se": "single_pilot_se"}, uid, bid
            )
            assert len(result.parse_warnings) == 1


# ── Route: import rollback ────────────────────────────────────────────────────


class TestImportRollback:
    def test_rollback_removes_entries(self, app, client):
        with app.app_context():
            uid = _make_user("rollback@example.com")
        _login(client, uid)

        with app.app_context():
            from datetime import datetime, timezone

            batch = LogbookImportBatch(
                pilot_user_id=uid,
                source_filename="rb.csv",
                imported_at=datetime.now(timezone.utc),
                row_count=2,
            )
            db.session.add(batch)
            db.session.flush()
            for d in [date(2024, 1, 1), date(2024, 1, 2)]:
                db.session.add(
                    PilotLogbookEntry(
                        pilot_user_id=uid,
                        import_batch_id=batch.id,
                        source="import",
                        date=d,
                    )
                )
            db.session.commit()
            bid = batch.id

        rv = client.post(f"/pilot/logbook/import/{bid}/rollback")
        assert rv.status_code in (302, 200)

        with app.app_context():
            assert PilotLogbookEntry.query.filter_by(import_batch_id=bid).count() == 0
            assert db.session.get(LogbookImportBatch, bid) is None

    def test_rollback_wrong_user_404(self, app, client):
        with app.app_context():
            uid = _make_user("rb2@example.com")
            other_uid = _make_user("other@example.com")
        _login(client, uid)

        with app.app_context():
            from datetime import datetime, timezone

            batch = LogbookImportBatch(
                pilot_user_id=other_uid,
                source_filename="rb.csv",
                imported_at=datetime.now(timezone.utc),
            )
            db.session.add(batch)
            db.session.commit()
            bid = batch.id

        rv = client.post(f"/pilot/logbook/import/{bid}/rollback")
        assert rv.status_code == 404


# ── Route: import upload (GET) ────────────────────────────────────────────────


class TestImportUploadRoute:
    def test_get_shows_upload_form(self, app, client):
        with app.app_context():
            uid = _make_user("upload1@example.com")
        _login(client, uid)

        rv = client.get("/pilot/logbook/import")
        assert rv.status_code == 200
        assert b"logbook_file" in rv.data

    def test_post_no_file_returns_422(self, app, client):
        with app.app_context():
            uid = _make_user("upload2@example.com")
        _login(client, uid)

        rv = client.post("/pilot/logbook/import", data={})
        assert rv.status_code == 422

    def test_post_unsupported_extension_returns_422(self, app, client):
        with app.app_context():
            uid = _make_user("upload3@example.com")
        _login(client, uid)

        rv = client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(b"data"), "log.flightlog")},
            content_type="multipart/form-data",
        )
        assert rv.status_code == 422

    def test_post_valid_csv_shows_mapping_page(self, app, client):
        with app.app_context():
            uid = _make_user("upload4@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        rv = client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(csv_data), "log.csv")},
            content_type="multipart/form-data",
        )
        assert rv.status_code == 200
        assert b"mapping_" in rv.data  # mapping dropdowns present


# ── Service: additional parser edge cases ─────────────────────────────────────


class TestParserEdgeCases:
    def test_is_numeric_str_returns_true(self):
        from pilots.logbook_import import _is_numeric_str  # pyright: ignore[reportMissingImports]

        assert _is_numeric_str("3.14")
        assert _is_numeric_str("42")
        assert not _is_numeric_str("abc")

    def test_parse_date_int_returns_none(self):
        assert parse_date_value(12345) is None

    def test_parse_date_empty_string_returns_none(self):
        assert parse_date_value("") is None

    def test_parse_date_date_object(self):
        assert parse_date_value(date(2024, 5, 1)) == date(2024, 5, 1)

    def test_parse_time_from_datetime(self):
        from datetime import datetime

        assert parse_time_value(datetime(2024, 1, 1, 9, 30)) == time(9, 30)

    def test_parse_time_non_string_non_datetime_returns_none(self):
        assert parse_time_value(12345) is None

    def test_parse_time_invalid_hhmm_values(self):
        # Values that match pattern but are invalid time (e.g. 99:99)
        assert parse_time_value("99:99") is None

    def test_parse_duration_from_datetime(self):
        from datetime import datetime

        dt = datetime(2024, 1, 1, 1, 24)
        assert parse_duration_value(dt) == 1.4

    def test_parse_duration_negative_float_returns_none(self):
        assert parse_duration_value(-1.0) is None

    def test_parse_duration_invalid_string_returns_none(self):
        assert parse_duration_value("not-a-number") is None

    def test_parse_duration_infinite_string_returns_none(self):
        # float("inf") parses without raising and passes a naive `>= 0`
        # sign check (inf is not negative), so this needed an explicit
        # isfinite() guard.
        assert parse_duration_value("inf") is None
        assert parse_duration_value("Infinity") is None

    def test_parse_duration_infinite_float_returns_none(self):
        assert parse_duration_value(float("inf")) is None

    def test_parse_duration_nan_string_returns_none(self):
        assert parse_duration_value("nan") is None

    def test_parse_int_value_float(self):
        from pilots.logbook_import import parse_int_value  # pyright: ignore[reportMissingImports]

        assert parse_int_value(2.0) == 2
        assert parse_int_value(-1.0) is None

    def test_parse_int_value_str(self):
        from pilots.logbook_import import parse_int_value  # pyright: ignore[reportMissingImports]

        assert parse_int_value("3") == 3
        assert parse_int_value("") is None
        assert parse_int_value("bad") is None

    def test_parse_int_value_overflow_string_returns_none(self):
        # float("1e400") overflows to inf, and int(inf) raises OverflowError
        # rather than ValueError — found by the fuzz_logbook_value_parsers
        # harness fuzzing an uploaded landings_day/landings_night cell.
        from pilots.logbook_import import parse_int_value  # pyright: ignore[reportMissingImports]

        assert parse_int_value("1e400") is None
        assert parse_int_value("-1e400") is None

    def test_parse_int_value_overflow_float_returns_none(self):
        from pilots.logbook_import import parse_int_value  # pyright: ignore[reportMissingImports]

        assert parse_int_value(float("inf")) is None

    def test_parse_int_value_negative_string_returns_none(self):
        # The string branch parsed "-2" via int(float("-2")) without the
        # negative-value guard the int/float branches already had — found by
        # the fuzz_logbook_value_parsers harness. A negative landings_day/
        # landings_night cell must be rejected like any other invalid value,
        # not silently stored as a negative count.
        from pilots.logbook_import import parse_int_value  # pyright: ignore[reportMissingImports]

        assert parse_int_value("-2") is None
        assert parse_int_value("-2.5") is None

    def test_subtotal_date_idx_beyond_row_length(self):
        # date_col_idx is beyond the row length
        assert _is_subtotal_row(["x", "y"], date_col_idx=5)


class TestExcelParsing:
    def _make_xlsx(self, rows: list[list]) -> bytes:
        import openpyxl  # pyright: ignore[reportMissingImports]

        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_xlsx_parsed_correctly(self):
        data = self._make_xlsx(
            [
                ["Date", "From", "To", "SE", "PIC"],
                ["15/03/24", "EBNM", "EBAW", "0.5", "0.5"],
            ]
        )
        pf = parse_file(data, "log.xlsx")
        assert pf.header_row_index == 0
        assert "date" in pf.norm_cols
        assert len(pf.data_rows) == 1

    def test_invalid_xlsx_raises(self):
        with pytest.raises(ValueError, match="Could not open"):
            parse_file(b"not an xlsx file", "log.xlsx")


class TestPreviewRows:
    def _make_parsed(self, data_rows, cols=None) -> ParsedFile:
        if cols is None:
            cols = ["date", "from", "to", "se", "ignored_col"]
        return ParsedFile(
            norm_cols=cols,
            raw_cols=cols,
            header_row_index=0,
            data_rows=data_rows,
            fingerprint=_fingerprint(cols),
        )

    def test_preview_skips_subtotal_rows(self):
        rows = [
            [date(2024, 1, 1), "EBNM", "EBAW", "0.5", "x"],
            [timedelta(hours=10), "sub", "sub", "10", "x"],  # subtotal
            [date(2024, 1, 2), "EBAW", "EBNM", "0.8", "x"],
        ]
        mapping = {
            "date": "date",
            "from": "departure_place",
            "to": "arrival_place",
            "se": "single_pilot_se",
            "ignored_col": "ignore",
        }
        pf = self._make_parsed(rows)
        result = preview_rows(pf, mapping, n=5)
        assert len(result) == 2
        assert all("departure_place" in r for r in result)

    def test_preview_skips_ignored_cols(self):
        rows = [[date(2024, 1, 1), "EBNM", "EBAW", "0.5", "should_not_appear"]]
        mapping = {
            "date": "date",
            "from": "departure_place",
            "to": "arrival_place",
            "se": "single_pilot_se",
            "ignored_col": "ignore",
        }
        pf = self._make_parsed(rows)
        result = preview_rows(pf, mapping, n=5)
        assert len(result) == 1
        assert "ignored_col" not in result[0]

    def test_preview_handles_short_row(self):
        # Row is shorter than the number of columns
        rows = [[date(2024, 1, 1), "EBNM"]]  # missing cols 2-4
        mapping = {
            "date": "date",
            "from": "departure_place",
            "to": "arrival_place",
            "se": "single_pilot_se",
            "ignored_col": "ignore",
        }
        pf = self._make_parsed(rows)
        result = preview_rows(pf, mapping, n=5)
        assert len(result) == 1
        assert result[0]["arrival_place"] is None

    def test_preview_respects_n_limit(self):
        rows = [[date(2024, 1, i), "A", "B", "0.5", "x"] for i in range(1, 11)]
        mapping = {
            "date": "date",
            "from": "departure_place",
            "to": "arrival_place",
            "se": "single_pilot_se",
            "ignored_col": "ignore",
        }
        pf = self._make_parsed(rows)
        result = preview_rows(pf, mapping, n=3)
        assert len(result) == 3


class TestFuzzyMappingEdgeCases:
    def _make_parsed(self, cols: list[str]) -> ParsedFile:
        return ParsedFile(
            norm_cols=cols,
            raw_cols=cols,
            header_row_index=0,
            data_rows=[],
            fingerprint=_fingerprint(cols),
        )

    def test_empty_saved_columns_skipped(self, app):
        pf = self._make_parsed(["date", "from"])

        class FakeMapping:
            source_fingerprint = "other"
            column_mapping = json.dumps({"date": "date"})
            source_columns = json.dumps([])  # empty list — should be skipped
            id = 1

        proposal = propose_mapping(pf, [FakeMapping()])
        # Should fall back to alias since no valid saved cols
        assert proposal.match_type == "alias"


# ── Route: full execute flow ──────────────────────────────────────────────────


class TestImportExecuteRoute:
    """Integration tests for the upload → execute flow."""

    def _upload_csv(self, client, csv_bytes: bytes, filename: str = "log.csv"):
        return client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(csv_bytes), filename)},
            content_type="multipart/form-data",
        )

    def _execute(self, client, extra_form: dict | None = None):
        form = {
            "mapping_date": "date",
            "mapping_from": "departure_place",
            "mapping_to": "arrival_place",
            "mapping_se": "single_pilot_se",
            "mapping_pic": "function_pic",
        }
        if extra_form:
            form.update(extra_form)
        return client.post("/pilot/logbook/import/execute", data=form)

    def test_execute_no_session_redirects(self, app, client):
        with app.app_context():
            uid = _make_user("ex_nosession@example.com")
        _login(client, uid)

        rv = client.post("/pilot/logbook/import/execute", data={})
        assert rv.status_code == 302
        assert "/import" in rv.headers["Location"]

    def test_execute_success_redirects_to_history(self, app, client):
        with app.app_context():
            uid = _make_user("ex_success@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n16/03/24,EBAW,EBNM,0.8,0.8\n"
        rv1 = self._upload_csv(client, csv_data)
        assert rv1.status_code == 200

        rv2 = self._execute(client)
        assert rv2.status_code == 302
        assert "history" in rv2.headers["Location"]

        with app.app_context():
            count = PilotLogbookEntry.query.filter_by(pilot_user_id=uid).count()
            assert count == 2

    def test_execute_no_date_mapped_returns_422(self, app, client):
        with app.app_context():
            uid = _make_user("ex_nodate@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        self._upload_csv(client, csv_data)

        # Submit with date mapped to ignore
        rv = client.post(
            "/pilot/logbook/import/execute",
            data={
                "mapping_date": "ignore",
                "mapping_from": "departure_place",
                "mapping_to": "arrival_place",
                "mapping_se": "single_pilot_se",
                "mapping_pic": "function_pic",
            },
        )
        assert rv.status_code == 422

    def test_execute_with_opening_balance(self, app, client):
        with app.app_context():
            uid = _make_user("ex_ob@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        self._upload_csv(client, csv_data)

        rv = self._execute(
            client, {"ob_single_pilot_se": "50.0", "ob_function_pic": "50.0"}
        )
        assert rv.status_code == 302

        with app.app_context():
            entries = PilotLogbookEntry.query.filter_by(pilot_user_id=uid).all()
            # 1 real + 1 opening balance
            assert len(entries) == 2
            ob = next(e for e in entries if e.remarks == "Opening balance (imported)")
            assert float(ob.single_pilot_se) == 50.0

    def test_execute_saves_mapping_for_second_import(self, app, client):
        with app.app_context():
            uid = _make_user("ex_remap@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        # First import
        self._upload_csv(client, csv_data)
        self._execute(client)

        # Second import — should recognise saved mapping (exact fingerprint)
        rv1 = self._upload_csv(client, csv_data)
        assert rv1.status_code == 200
        assert b"Recognised format" in rv1.data or b"recognised" in rv1.data.lower()

    def test_import_history_shows_batches(self, app, client):
        with app.app_context():
            uid = _make_user("history1@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        self._upload_csv(client, csv_data)
        self._execute(client)

        rv = client.get("/pilot/logbook/import/history")
        assert rv.status_code == 200
        assert b"log.csv" in rv.data

    def test_execute_opening_balance_with_no_entries(self, app, client):
        """Opening balance when all rows are subtotals/skipped uses today as anchor."""
        with app.app_context():
            uid = _make_user("ex_ob_empty@example.com")
        _login(client, uid)

        # CSV where every data row is a subtotal (TOTAL in date field)
        csv_data = b"Date,From,To,SE,PIC\nTOTAL,,, 10.0, 10.0\n"
        self._upload_csv(client, csv_data)

        rv = self._execute(client, {"ob_single_pilot_se": "10.0"})
        assert rv.status_code == 302

        with app.app_context():
            entries = PilotLogbookEntry.query.filter_by(pilot_user_id=uid).all()
            ob_entries = [
                e for e in entries if e.remarks == "Opening balance (imported)"
            ]
            assert len(ob_entries) == 1

    def test_execute_with_departure_and_landings_columns(self, app, client):
        """Cover departure_time (line 531) and landings_day (line 539) branches."""
        with app.app_context():
            uid = _make_user("ex_timelanding@example.com")
        _login(client, uid)

        csv_data = (
            b"Date,From,Time,To,Time,DAY,NIGHT\n15/03/24,EBNM,09:00,EBAW,10:30,1,0\n"
        )
        self._upload_csv(client, csv_data)

        rv = client.post(
            "/pilot/logbook/import/execute",
            data={
                "mapping_date": "date",
                "mapping_from": "departure_place",
                "mapping_time": "departure_time",
                "mapping_to": "arrival_place",
                "mapping_time_2": "arrival_time",
                "mapping_day": "landings_day",
                "mapping_night": "landings_night",
            },
        )
        assert rv.status_code == 302

        with app.app_context():
            e = PilotLogbookEntry.query.filter_by(pilot_user_id=uid).first()
            assert e is not None
            assert e.departure_time == time(9, 0)
            assert e.landings_day == 1

    def test_execute_with_skipped_rows_shows_warning(self, app, client):
        """Cover the skipped-rows flash (routes line ~645)."""
        with app.app_context():
            uid = _make_user("ex_skipped@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\nbaddate,EBNM,EBAW,0.5,0.5\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        self._upload_csv(client, csv_data)
        rv = self._execute(client)
        # Redirect to history
        assert rv.status_code == 302

        with app.app_context():
            batch = LogbookImportBatch.query.filter_by(pilot_user_id=uid).first()
            assert batch is not None
            assert batch.skipped_count == 1
            assert batch.row_count == 1

    def test_execute_updates_existing_mapping(self, app, client):
        """Cover the 'update saved mapping' branch (routes line ~586-590)."""
        with app.app_context():
            uid = _make_user("ex_updatemap@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        # First import — creates LogbookImportMapping
        self._upload_csv(client, csv_data)
        self._execute(client)

        # Second import with same file — existing mapping updated
        self._upload_csv(client, csv_data)
        rv = self._execute(client)
        assert rv.status_code == 302

        with app.app_context():
            mappings = LogbookImportMapping.query.filter_by(pilot_user_id=uid).all()
            # Should reuse the same mapping record (not create two)
            assert len(mappings) == 1

    def test_execute_tmp_file_not_found_redirects(self, app, client):
        """Cover the 'tmp file not found' path (routes line ~526-528)."""
        with app.app_context():
            uid = _make_user("ex_notmp@example.com")
        _login(client, uid)

        # Manually set a session with a non-existent tmp_path
        with client.session_transaction() as sess:
            sess["logbook_import"] = {
                "uid": uid,
                "tmp_path": "/nonexistent/file.csv",
                "original_filename": "file.csv",
                "norm_cols": ["date", "from"],
                "raw_cols": ["Date", "From"],
                "fingerprint": "abc123",
            }

        rv = client.post(
            "/pilot/logbook/import/execute",
            data={"mapping_date": "date", "mapping_from": "departure_place"},
        )
        assert rv.status_code == 302
        assert "/import" in rv.headers["Location"]

    def test_execute_total_mismatch_warning_flash(self, app, client):
        """Cover routes.py total_mismatch_warnings flash block (lines 959-971)."""
        with app.app_context():
            uid = _make_user("tm_flash@example.com")
        _login(client, uid)

        # SE=0.5 but Total Flight Time=1.5 — mismatch triggers warning
        csv_data = (
            b"Date,From,To,SE,PIC,Total Flight Time\n15/03/24,EBNM,EBAW,0.5,0.5,1.5\n"
        )
        self._upload_csv(client, csv_data)
        rv = self._execute(
            client,
            extra_form={"mapping_total flight time": "total_flight_time_check"},
        )
        assert rv.status_code == 302

        with client.session_transaction() as sess:
            flashes = sess.get("_flashes", [])
        warning_messages = [msg for cat, msg in flashes if cat == "warning"]
        assert any("total flight time" in m.lower() for m in warning_messages)

    def test_execute_total_mismatch_warning_flash_n_gt_3(self, app, client):
        """Cover the n>3 ellipsis branch of total_mismatch_warnings (line 969-970)."""
        with app.app_context():
            uid = _make_user("tm_flash4@example.com")
        _login(client, uid)

        # Four rows where total (2.0) doesn't match SE (0.5)
        csv_data = (
            b"Date,From,To,SE,PIC,Total Flight Time\n"
            b"15/03/24,EBNM,EBAW,0.5,0.5,2.0\n"
            b"16/03/24,EBNM,EBAW,0.5,0.5,2.0\n"
            b"17/03/24,EBNM,EBAW,0.5,0.5,2.0\n"
            b"18/03/24,EBNM,EBAW,0.5,0.5,2.0\n"
        )
        self._upload_csv(client, csv_data)
        rv = self._execute(
            client,
            extra_form={"mapping_total flight time": "total_flight_time_check"},
        )
        assert rv.status_code == 302

        with client.session_transaction() as sess:
            flashes = sess.get("_flashes", [])
        warning_messages = [msg for cat, msg in flashes if cat == "warning"]
        assert any("+1" in m for m in warning_messages)

    def test_execute_parse_warnings_flash(self, app, client):
        """Cover routes.py parse_warnings flash block (lines 667-674), including n>3 branch."""
        with app.app_context():
            uid = _make_user("pw_flash@example.com")
        _login(client, uid)

        # Four rows with bad "se" values — triggers the n > 3 ellipsis branch
        csv_data = (
            b"Date,From,To,SE,PIC\n"
            b"15/03/24,EBNM,EBAW,bad1,0.5\n"
            b"16/03/24,EBNM,EBAW,bad2,0.5\n"
            b"17/03/24,EBNM,EBAW,bad3,0.5\n"
            b"18/03/24,EBNM,EBAW,bad4,0.5\n"
        )
        self._upload_csv(client, csv_data)

        rv = self._execute(client)
        assert rv.status_code == 302

        with client.session_transaction() as sess:
            flashes = sess.get("_flashes", [])
        warning_messages = [msg for cat, msg in flashes if cat == "warning"]
        assert any("could not be parsed" in m for m in warning_messages)
        assert any("+1" in m for m in warning_messages)


# ── Service: CSV edge cases ───────────────────────────────────────────────────


class TestCSVEdgeCases:
    def test_latin1_csv_decoded(self):
        """Cover the latin-1 fallback (line 234-235)."""
        # é and è are valid latin-1 but invalid UTF-8 — need ≥4 columns to pass header detection
        row = b"Date,From,To,SE,PIC\r\n15/03/24,\xe9l\xe8ve,EBAW,0.5,0.5\r\n"
        pf = parse_file(row, "log.csv")
        assert len(pf.data_rows) == 1
        assert "date" in pf.norm_cols

    def test_csv_sniffer_fallback(self):
        """Cover the csv.Sniffer fallback (lines 240-241) using mock."""
        import csv as _csv
        from unittest.mock import patch

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        with patch.object(
            _csv.Sniffer, "sniff", side_effect=_csv.Error("cant determine")
        ):
            pf = parse_file(csv_data, "log.csv")
        assert "date" in pf.norm_cols

    def test_csv_error_from_reader_raises_value_error(self):
        # An embedded newline inside an unquoted field with a sniffed dialect
        # that doesn't expect it makes the stdlib csv reader itself raise
        # csv.Error — found by the fuzz_logbook_parse_file harness fuzzing
        # raw uploaded CSV bytes. Must surface as the documented ValueError,
        # not an unhandled 500.
        data = bytes.fromhex("3d000000000000000000000d000000")
        with pytest.raises(ValueError, match="Could not parse CSV file"):
            parse_file(data, "log.csv")


class TestParseIntEdgeCases:
    def test_negative_int_returns_none(self):
        from pilots.logbook_import import parse_int_value  # pyright: ignore[reportMissingImports]

        assert parse_int_value(-1) is None

    def test_none_returns_none(self):
        from pilots.logbook_import import parse_int_value  # pyright: ignore[reportMissingImports]

        assert parse_int_value(None) is None  # type: ignore[arg-type]


# ── Route: remaining edge cases ───────────────────────────────────────────────


class TestImportRouteEdgeCases:
    def test_upload_file_too_large_returns_422(self, app, client):
        """Cover file-too-large branch (routes line 465-466)."""
        with app.app_context():
            uid = _make_user("toolarge@example.com")
        _login(client, uid)

        big_data = b"x" * (11 * 1024 * 1024)  # 11 MB
        rv = client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(big_data), "big.csv")},
            content_type="multipart/form-data",
        )
        assert rv.status_code == 422

    def test_upload_unparseable_csv_returns_422(self, app, client):
        """Cover parse_file ValueError in import_upload (routes line 470-472)."""
        with app.app_context():
            uid = _make_user("badcsv@example.com")
        _login(client, uid)

        # All-numeric rows → header detection fails → ValueError
        bad_csv = b"1.0,2.0,3.0,4.0,5.0\n6.0,7.0,8.0,9.0,10.0\n"
        rv = client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(bad_csv), "bad.csv")},
            content_type="multipart/form-data",
        )
        assert rv.status_code == 422

    def test_upload_twice_cleans_previous_tmp(self, app, client):
        """Cover _cleanup_previous_tmp when a tmp file exists (routes line 431-436)."""
        with app.app_context():
            uid = _make_user("twoupload@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        # First upload — creates tmp file
        rv1 = client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(csv_data), "log.csv")},
            content_type="multipart/form-data",
        )
        assert rv1.status_code == 200

        # Get the tmp path from session
        with client.session_transaction() as sess:
            first_tmp = sess.get("logbook_import", {}).get("tmp_path")
        assert first_tmp and os.path.isfile(first_tmp)

        # Second upload — should delete the first tmp
        rv2 = client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(csv_data), "log2.csv")},
            content_type="multipart/form-data",
        )
        assert rv2.status_code == 200
        # Previous tmp should be gone
        assert not os.path.isfile(first_tmp)

    def test_cleanup_oserror_is_silenced(self, app, client):
        """Cover OSError branch in _cleanup_previous_tmp (routes lines 438-439)."""
        with app.app_context():
            uid = _make_user("cleanup_oserror@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        # First upload creates a tmp file in session
        rv1 = client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(csv_data), "log.csv")},
            content_type="multipart/form-data",
        )
        assert rv1.status_code == 200

        # Second upload triggers cleanup; os.remove raises OSError — must not crash
        from unittest.mock import patch

        with patch("pilots.routes.os.remove", side_effect=OSError("locked")):
            rv2 = client.post(
                "/pilot/logbook/import",
                data={"logbook_file": (io.BytesIO(csv_data), "log2.csv")},
                content_type="multipart/form-data",
            )
        assert rv2.status_code == 200

    def test_execute_reparse_fails_on_no_date_mapped_redirects(self, app, client):
        """Cover lines 545-547: re-parse fails when no date mapped."""
        with app.app_context():
            uid = _make_user("reparse_fail@example.com")
        _login(client, uid)

        # Set session pointing to a file with invalid content (no header)
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            f.write(b"1.0,2.0,3.0,4.0,5.0\n6.0,7.0,8.0,9.0,10.0\n")
            tmp = f.name

        with client.session_transaction() as sess:
            sess["logbook_import"] = {
                "uid": uid,
                "tmp_path": tmp,
                "original_filename": "bad.csv",
                "norm_cols": ["date", "from"],
                "raw_cols": ["Date", "From"],
                "fingerprint": "abc",
            }

        # Submit with no date column → triggers re-parse → parse fails → redirect
        rv = client.post(
            "/pilot/logbook/import/execute",
            data={"mapping_date": "ignore", "mapping_from": "departure_place"},
        )
        assert rv.status_code == 302
        assert "/import" in rv.headers["Location"]

        if os.path.isfile(tmp):
            os.remove(tmp)

    def test_execute_main_reparse_fails_redirects(self, app, client):
        """Cover lines 577-580: re-parse fails in main execute path."""
        with app.app_context():
            uid = _make_user("main_reparse_fail@example.com")
        _login(client, uid)

        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            f.write(b"1.0,2.0,3.0,4.0,5.0\n6.0,7.0,8.0,9.0,10.0\n")
            tmp = f.name

        with client.session_transaction() as sess:
            sess["logbook_import"] = {
                "uid": uid,
                "tmp_path": tmp,
                "original_filename": "bad.csv",
                "norm_cols": ["date", "from"],
                "raw_cols": ["Date", "From"],
                "fingerprint": "abc",
            }

        # Submit WITH date mapped → skips no-date validation → hits main reparse → fails
        rv = client.post(
            "/pilot/logbook/import/execute",
            data={"mapping_date": "date", "mapping_from": "departure_place"},
        )
        assert rv.status_code == 302
        assert "/import" in rv.headers["Location"]

        if os.path.isfile(tmp):
            os.remove(tmp)

    def test_execute_more_than_5_skipped_rows(self, app, client):
        """Cover the '… and X more' suffix (routes line 647)."""
        with app.app_context():
            uid = _make_user("manyskipped@example.com")
        _login(client, uid)

        # 6 rows with bad dates + 1 good row
        bad_rows = "\n".join(f"baddate{i},EBNM,EBAW,0.5,0.5" for i in range(6))
        csv_data = (
            f"Date,From,To,SE,PIC\n{bad_rows}\n15/03/24,EBNM,EBAW,0.5,0.5\n".encode()
        )

        client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(csv_data), "log.csv")},
            content_type="multipart/form-data",
        )
        rv = client.post(
            "/pilot/logbook/import/execute",
            data={
                "mapping_date": "date",
                "mapping_from": "departure_place",
                "mapping_to": "arrival_place",
                "mapping_se": "single_pilot_se",
                "mapping_pic": "function_pic",
            },
        )
        assert rv.status_code == 302

        with app.app_context():
            batch = LogbookImportBatch.query.filter_by(pilot_user_id=uid).first()
            assert batch is not None
            assert batch.skipped_count == 6

    def test_execute_osremove_oserror_handled(self, app, client):
        """Cover OSError on os.remove(tmp_path) (routes line 630-631)."""
        from unittest.mock import patch

        with app.app_context():
            uid = _make_user("removeerr@example.com")
        _login(client, uid)

        csv_data = b"Date,From,To,SE,PIC\n15/03/24,EBNM,EBAW,0.5,0.5\n"
        client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(csv_data), "log.csv")},
            content_type="multipart/form-data",
        )

        with patch("os.remove", side_effect=OSError("disk busy")):
            rv = client.post(
                "/pilot/logbook/import/execute",
                data={
                    "mapping_date": "date",
                    "mapping_from": "departure_place",
                    "mapping_to": "arrival_place",
                    "mapping_se": "single_pilot_se",
                    "mapping_pic": "function_pic",
                },
            )
        # Even with OSError on remove, import should succeed
        assert rv.status_code == 302
        assert "history" in rv.headers["Location"]


# ── Group-header detection ────────────────────────────────────────────────────


def _make_xlsx_with_merges(
    group_row: list[tuple[str, int]],  # [(label, span), ...]
    header_row: list[str],
    data_rows: list[list] | None = None,
) -> bytes:
    """Build an xlsx with a merged group-header row followed by a column-name row.

    *group_row* is a list of (label, span) pairs — each label occupies *span*
    consecutive columns; an empty label ("") writes no value but still advances
    the column pointer.
    """
    import openpyxl  # pyright: ignore[reportMissingImports]

    wb = openpyxl.Workbook()
    ws = wb.active

    # Write group row with merged cells
    col = 1
    for label, span in group_row:
        if label:
            ws.cell(row=1, column=col, value=label)
            if span > 1:
                ws.merge_cells(
                    start_row=1,
                    start_column=col,
                    end_row=1,
                    end_column=col + span - 1,
                )
        col += span

    # Write header row
    for j, name in enumerate(header_row, start=1):
        ws.cell(row=2, column=j, value=name)

    for row in data_rows or []:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestMergeLabelMap:
    def _ws_with_merges(
        self, group_row: list[tuple[str, int]]
    ):  # returns openpyxl worksheet
        import openpyxl  # pyright: ignore[reportMissingImports]

        wb = openpyxl.Workbook()
        ws = wb.active
        col = 1
        for label, span in group_row:
            if label:
                ws.cell(row=1, column=col, value=label)
                if span > 1:
                    ws.merge_cells(
                        start_row=1,
                        start_column=col,
                        end_row=1,
                        end_column=col + span - 1,
                    )
            col += span
        return ws

    def test_single_merged_region(self):
        ws = self._ws_with_merges([("DEPARTURE & ARRIVAL", 4)])
        m = _merge_label_map(ws)
        assert m[(0, 0)] == "DEPARTURE & ARRIVAL"
        assert m[(0, 1)] == "DEPARTURE & ARRIVAL"
        assert m[(0, 2)] == "DEPARTURE & ARRIVAL"
        assert m[(0, 3)] == "DEPARTURE & ARRIVAL"
        assert (0, 4) not in m

    def test_two_disjoint_merged_regions(self):
        ws = self._ws_with_merges([("GROUP A", 2), ("GROUP B", 3)])
        m = _merge_label_map(ws)
        assert m[(0, 0)] == "GROUP A"
        assert m[(0, 1)] == "GROUP A"
        assert m[(0, 2)] == "GROUP B"
        assert m[(0, 3)] == "GROUP B"
        assert m[(0, 4)] == "GROUP B"

    def test_empty_label_merge_excluded(self):
        ws = self._ws_with_merges([("", 3), ("LANDINGS", 2)])
        m = _merge_label_map(ws)
        # Empty-label region excluded; only LANDINGS present (cols 3-4, 0-based)
        assert (0, 0) not in m
        assert (0, 1) not in m
        assert (0, 2) not in m
        assert m[(0, 3)] == "LANDINGS"
        assert m[(0, 4)] == "LANDINGS"

    def test_no_merged_cells_returns_empty(self):
        import openpyxl  # pyright: ignore[reportMissingImports]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "DATE"
        ws["B1"] = "FROM"
        m = _merge_label_map(ws)
        assert m == {}

    def test_merged_region_with_none_value_excluded(self):
        """Merged region whose top-left cell is None → skipped (line 295)."""
        import openpyxl  # pyright: ignore[reportMissingImports]

        wb = openpyxl.Workbook()
        ws = wb.active
        # A1:C1 merged but left empty (None value) → must be excluded
        ws.merge_cells("A1:C1")
        # D1:E1 merged with a label → must be included
        ws["D1"] = "LANDINGS"
        ws.merge_cells("D1:E1")

        m = _merge_label_map(ws)
        assert (0, 0) not in m
        assert (0, 1) not in m
        assert (0, 2) not in m
        assert m[(0, 3)] == "LANDINGS"
        assert m[(0, 4)] == "LANDINGS"


class TestGroupLabelsFromMap:
    def test_returns_labels_for_known_cols(self):
        m = {(1, 0): "GRP A", (1, 1): "GRP A", (1, 2): "GRP B"}
        labels = _group_labels_from_map(m, row_idx=1, width=4)
        assert labels == ["GRP A", "GRP A", "GRP B", ""]

    def test_empty_map_returns_all_empty(self):
        labels = _group_labels_from_map({}, row_idx=0, width=3)
        assert labels == ["", "", ""]

    def test_width_zero_returns_empty_list(self):
        assert _group_labels_from_map({(0, 0): "X"}, row_idx=0, width=0) == []


class TestGroupLabelsHeuristic:
    def test_sparse_row_forward_filled(self):
        row = ["DEPARTURE & ARRIVAL", None, None, None, "LANDINGS", None]
        labels = _group_labels_heuristic(row, width=6)
        assert labels == [
            "DEPARTURE & ARRIVAL",
            "DEPARTURE & ARRIVAL",
            "DEPARTURE & ARRIVAL",
            "DEPARTURE & ARRIVAL",
            "LANDINGS",
            "LANDINGS",
        ]

    def test_no_span_returns_none(self):
        # Every cell immediately follows the previous — no gap
        row = ["DATE", "FROM", "TO", "SE"]
        assert _group_labels_heuristic(row, width=4) is None

    def test_single_nonempty_returns_none(self):
        row = ["TITLE", None, None, None]
        assert _group_labels_heuristic(row, width=4) is None

    def test_empty_row_returns_none(self):
        assert _group_labels_heuristic([None, None, None], width=3) is None

    def test_width_wider_than_row_pads_with_last(self):
        row = ["GRP A", None, "GRP B"]
        labels = _group_labels_heuristic(row, width=5)
        assert labels is not None
        assert labels[0] == "GRP A"
        assert labels[2] == "GRP B"
        # Padding beyond row end continues last value
        assert labels[3] == "GRP B"
        assert labels[4] == "GRP B"

    def test_two_adjacent_values_not_a_span(self):
        # "A" at col 0 and "B" at col 1 — no gap, not a group row
        row = ["A", "B", None, None]
        assert _group_labels_heuristic(row, width=4) is None


class TestApplyGroupLabels:
    def test_prepends_label_to_col(self):
        result = _apply_group_labels(["GRP A", "GRP A", ""], ["FROM", "TIME", "DATE"])
        assert result == ["GRP A FROM", "GRP A TIME", "DATE"]

    def test_empty_label_leaves_col_unchanged(self):
        result = _apply_group_labels(["", "GRP"], ["DATE", "FROM"])
        assert result == ["DATE", "GRP FROM"]

    def test_empty_col_uses_label_only(self):
        result = _apply_group_labels(["GRP", ""], ["", "DATE"])
        assert result == ["GRP", "DATE"]

    def test_both_empty_produces_empty_string(self):
        result = _apply_group_labels([""], [""])
        assert result == [""]

    def test_group_labels_shorter_than_header(self):
        # Extra header columns beyond group_labels length get no prefix
        result = _apply_group_labels(["GRP"], ["FROM", "TIME", "TO"])
        assert result == ["GRP FROM", "TIME", "TO"]


class TestGroupHeaderExcelIntegration:
    """End-to-end: parse_file on an xlsx with merged group headers."""

    def test_easa_logbook_style_columns_resolved(self):
        data = _make_xlsx_with_merges(
            group_row=[
                ("DEPARTURE & ARRIVAL", 4),
                ("LANDINGS", 2),
                ("AIRCRAFT CATEGORY", 2),
            ],
            header_row=["FROM", "TIME", "TO", "TIME", "DAY", "NIGHT", "SE", "ME"],
            data_rows=[["EBNM", "09:00", "EBAW", "10:30", "1", "0", "0.5", "0.0"]],
        )
        pf = parse_file(data, "log.xlsx")
        assert pf.header_row_index == 1
        assert "departure & arrival from" in pf.norm_cols
        assert "departure & arrival time" in pf.norm_cols
        assert "departure & arrival time_2" in pf.norm_cols
        assert "departure & arrival to" in pf.norm_cols
        assert "landings day" in pf.norm_cols
        assert "landings night" in pf.norm_cols
        assert "aircraft category se" in pf.norm_cols
        assert "aircraft category me" in pf.norm_cols

    def test_no_group_row_parses_normally(self):
        import openpyxl  # pyright: ignore[reportMissingImports]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "From", "To", "SE", "PIC"])
        ws.append(["15/03/24", "EBNM", "EBAW", "0.5", "Smith"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        pf = parse_file(buf.read(), "log.xlsx")
        assert pf.header_row_index == 0
        assert pf.norm_cols == ["date", "from", "to", "se", "pic"]
        assert len(pf.data_rows) == 1

    def test_group_row_all_empty_ignored(self):
        """A row of entirely empty merges above the header is treated as no group row."""
        data = _make_xlsx_with_merges(
            group_row=[("", 4)],
            header_row=["Date", "From", "To", "SE"],
            data_rows=[],
        )
        pf = parse_file(data, "log.xlsx")
        # Group labels all empty → no prepending
        assert pf.norm_cols == ["date", "from", "to", "se"]

    def test_duplicate_prefixed_columns_disambiguated(self):
        """Two TIME columns under the same group span → disambiguated as _2."""
        data = _make_xlsx_with_merges(
            group_row=[("DEPARTURE & ARRIVAL", 4)],
            header_row=["FROM", "TIME", "TO", "TIME"],
            data_rows=[],
        )
        pf = parse_file(data, "log.xlsx")
        assert "departure & arrival time" in pf.norm_cols
        assert "departure & arrival time_2" in pf.norm_cols

    def test_second_load_exception_falls_back_gracefully(self):
        """Exception in the second load_workbook call is silenced (lines 385-386)."""
        from unittest.mock import patch

        import openpyxl  # pyright: ignore[reportMissingImports]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "From", "To", "SE", "PIC"])
        ws.append(["15/03/24", "EBNM", "EBAW", "0.5", "Smith"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.read()

        call_count: list[int] = [0]
        real_load = openpyxl.load_workbook

        def fail_on_second(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] == 2:
                raise RuntimeError("simulated second-load failure")
            return real_load(*args, **kwargs)

        with patch("pilots.logbook_import.openpyxl.load_workbook", fail_on_second):
            pf = parse_file(data, "log.xlsx")

        # Parsing still succeeds, just without group-label enrichment
        assert "date" in pf.norm_cols
        assert len(pf.data_rows) == 1

    def test_alias_mapping_resolves_group_prefixed_columns(self):
        from pilots.logbook_import import _alias_mapping  # pyright: ignore[reportMissingImports]

        data = _make_xlsx_with_merges(
            group_row=[("DEPARTURE & ARRIVAL", 4), ("LANDINGS", 2)],
            header_row=["FROM", "TIME", "TO", "TIME", "DAY", "NIGHT"],
            data_rows=[],
        )
        pf = parse_file(data, "log.xlsx")
        mapping = _alias_mapping(pf.norm_cols)
        assert mapping["departure & arrival from"] == "departure_place"
        assert mapping["departure & arrival time"] == "departure_time"
        assert mapping["departure & arrival time_2"] == "arrival_time"
        assert mapping["departure & arrival to"] == "arrival_place"
        assert mapping["landings day"] == "landings_day"
        assert mapping["landings night"] == "landings_night"


class TestGroupHeaderCsvHeuristic:
    """CSV files use the forward-fill heuristic (no merge metadata)."""

    def _make_csv(self, rows: list[list]) -> bytes:
        lines = [",".join(str(c) if c is not None else "" for c in row) for row in rows]
        return "\n".join(lines).encode()

    def test_csv_sparse_group_row_applied(self):
        data = self._make_csv(
            [
                # Sparse row: values at cols 0 and 4 with gaps → qualifies as group row
                ["DEPARTURE & ARRIVAL", "", "", "", "LANDINGS", ""],
                ["FROM", "TIME", "TO", "TIME", "DAY", "NIGHT"],
                ["EBNM", "09:00", "EBAW", "10:30", "1", "0"],
            ]
        )
        pf = parse_file(data, "log.csv")
        assert pf.header_row_index == 1
        assert "departure & arrival from" in pf.norm_cols
        assert "landings day" in pf.norm_cols
        assert "landings night" in pf.norm_cols

    def test_csv_dense_group_row_not_applied(self):
        """A dense row (no gaps) is not treated as a group row."""
        data = self._make_csv(
            [
                ["DATE", "FROM", "TO", "SE", "PIC"],
                ["15/03/24", "EBNM", "EBAW", "0.5", "Smith"],
            ]
        )
        pf = parse_file(data, "log.csv")
        assert pf.header_row_index == 0
        # No group prepending — column names unchanged
        assert "date" in pf.norm_cols
        assert "from" in pf.norm_cols


class TestPickBestExcelSheet:
    """_pick_best_excel_sheet selects the right sheet in multi-sheet workbooks."""

    def _make_multisheet_xlsx(
        self,
        sheets: list[tuple[str, list[list]]],
        active_idx: int = 0,
    ) -> bytes:
        import openpyxl  # pyright: ignore[reportMissingImports]

        wb = openpyxl.Workbook()
        for i, (name, rows) in enumerate(sheets):
            if i == 0:
                ws = wb.active
                ws.title = name  # type: ignore[union-attr]
            else:
                ws = wb.create_sheet(name)
            for row in rows:
                ws.append(row)  # type: ignore[union-attr]
        wb.active = wb.worksheets[active_idx]
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_preferred_name_wins_over_active_sheet(self):
        """Sheet named 'Logbook' is chosen even when 'Stats' is the active sheet."""
        data = self._make_multisheet_xlsx(
            sheets=[
                (
                    "Logbook",
                    [["Date", "From", "To", "SE"], ["15/03/24", "EBNM", "EBAW", "0.5"]],
                ),
                ("Stats", [["Total", "Day", "Night"], ["1:00", "0:30", "0:30"]]),
            ],
            active_idx=1,  # Stats is active
        )
        pf = parse_file(data, "log.xlsx")
        assert pf.header_row_index == 0
        assert "date" in pf.norm_cols
        assert "from" in pf.norm_cols
        assert len(pf.data_rows) == 1

    def test_preferred_name_case_insensitive(self):
        """Sheet named 'LOGBOOK' (upper-case) is still preferred."""
        data = self._make_multisheet_xlsx(
            sheets=[
                (
                    "LOGBOOK",
                    [["Date", "From", "To", "SE"], ["15/03/24", "EBNM", "EBAW", "0.5"]],
                ),
                ("Summary", [["Info"], ["x"]]),
            ],
            active_idx=1,
        )
        pf = parse_file(data, "log.xlsx")
        assert "date" in pf.norm_cols

    def test_score_fallback_picks_most_aliased_sheet(self):
        """When no preferred name, the sheet with more alias matches is chosen."""
        data = self._make_multisheet_xlsx(
            sheets=[
                ("Summary", [["Total", "Day", "Night"], ["1:00", "0:30", "0:30"]]),
                (
                    "Flights",
                    [
                        ["Date", "From", "To", "SE", "PIC"],
                        ["15/03/24", "EBNM", "EBAW", "0.5", "Smith"],
                    ],
                ),
            ],
            active_idx=0,  # Summary is active
        )
        pf = parse_file(data, "log.xlsx")
        # "Flights" matches preferred names list
        assert "date" in pf.norm_cols

    def test_single_sheet_workbook_unaffected(self):
        """A workbook with one sheet still works as before."""
        data = self._make_multisheet_xlsx(
            sheets=[
                (
                    "Sheet1",
                    [
                        ["Date", "From", "To", "SE", "PIC"],
                        ["15/03/24", "EBNM", "EBAW", "0.5", "Smith"],
                    ],
                )
            ],
        )
        pf = parse_file(data, "log.xlsx")
        assert "date" in pf.norm_cols
        assert len(pf.data_rows) == 1

    def test_merge_map_uses_correct_sheet(self):
        """The merge map is built from the selected sheet, not the active sheet.

        Workbook: 'Logbook' has merged group headers; 'Stats' is active and has none.
        After picking 'Logbook', group-prefixed column names must be present.
        """
        import openpyxl  # pyright: ignore[reportMissingImports]

        wb = openpyxl.Workbook()
        # Sheet 0: Logbook (with merged group header)
        ws_log = wb.active
        ws_log.title = "Logbook"  # type: ignore[union-attr]
        ws_log.append(["DEPARTURE & ARRIVAL", None, None, None])  # type: ignore[union-attr]
        ws_log.merge_cells("A1:D1")
        ws_log.append(["FROM", "TIME", "TO", "TIME"])  # type: ignore[union-attr]
        ws_log.append(["EBNM", "09:00", "EBAW", "10:30"])  # type: ignore[union-attr]
        # Sheet 1: Stats (active, no logbook columns)
        ws_stats = wb.create_sheet("Stats")
        ws_stats.append(["Total", "Day", "Night"])
        wb.active = wb.worksheets[1]
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        pf = parse_file(buf.read(), "log.xlsx")
        assert "departure & arrival from" in pf.norm_cols
        assert "departure & arrival time" in pf.norm_cols
        assert "departure & arrival to" in pf.norm_cols


# ── Import triggers aircraft log linking when managed aircraft matches ─────────


class TestImportExecuteAircraftLink:
    """Covers pilots/routes.py: db.session.commit() and flash when ac_created > 0."""

    def _upload_csv(self, client, csv_bytes: bytes):
        return client.post(
            "/pilot/logbook/import",
            data={"logbook_file": (io.BytesIO(csv_bytes), "log.csv")},
            content_type="multipart/form-data",
        )

    def _execute(self, client, extra: dict | None = None):
        form = {
            "mapping_date": "date",
            "mapping_from": "departure_place",
            "mapping_to": "arrival_place",
            "mapping_reg": "aircraft_registration",
            "mapping_se": "single_pilot_se",
            "mapping_pic": "function_pic",
        }
        if extra:
            form.update(extra)
        return client.post("/pilot/logbook/import/execute", data=form)

    def test_creates_flight_entry_and_flashes_info(self, app, client):
        with app.app_context():
            user = User(
                email="aclink_exec@example.com",
                password_hash=_pw_hash.hash("pw"),
                is_active=True,
            )
            db.session.add(user)
            db.session.flush()
            tenant = Tenant(name="AC Link Exec Hangar")
            db.session.add(tenant)
            db.session.flush()
            db.session.add(
                TenantUser(user_id=user.id, tenant_id=tenant.id, role=Role.OWNER)
            )
            ac = Aircraft(
                registration="OO-TST",
                tenant_id=tenant.id,
                make="Cessna",
                model="172S",
                flight_counter_offset=0.3,
            )
            db.session.add(ac)
            db.session.commit()
            uid = user.id
            ac_id = ac.id

        with client.session_transaction() as sess:
            sess["user_id"] = uid
            sess["pilot_access"] = True

        csv_data = b"Date,From,To,Reg,SE,PIC\n15/03/24,EBBR,EBOS,OO-TST,0.8,0.8\n"
        rv1 = self._upload_csv(client, csv_data)
        assert rv1.status_code == 200

        rv2 = self._execute(client)
        assert rv2.status_code == 302

        with app.app_context():
            flight = FlightEntry.query.filter_by(aircraft_id=ac_id).first()
            assert flight is not None
            assert flight.source == "logbook_import"

        with client.session_transaction() as sess:
            flashes = sess.get("_flashes", [])
        info_msgs = [msg for cat, msg in flashes if cat == "info"]
        assert any("aircraft log entr" in m for m in info_msgs)

    def test_does_not_link_to_another_tenants_aircraft(self, app, client):
        """A registration collision with another tenant's aircraft must not
        create a FlightEntry (and crew access) onto that other tenant's
        fleet — the match is scoped to the importing pilot's own tenant(s)."""
        with app.app_context():
            user = User(
                email="aclink_crosstenant@example.com",
                password_hash=_pw_hash.hash("pw"),
                is_active=True,
            )
            db.session.add(user)
            db.session.flush()
            own_tenant = Tenant(name="Own Hangar")
            db.session.add(own_tenant)
            db.session.flush()
            db.session.add(
                TenantUser(user_id=user.id, tenant_id=own_tenant.id, role=Role.OWNER)
            )
            other_tenant = Tenant(name="Other Hangar")
            db.session.add(other_tenant)
            db.session.flush()
            other_ac = Aircraft(
                registration="OO-TST",
                tenant_id=other_tenant.id,
                make="Cessna",
                model="172S",
                flight_counter_offset=0.3,
            )
            db.session.add(other_ac)
            db.session.commit()
            uid = user.id
            other_ac_id = other_ac.id

        with client.session_transaction() as sess:
            sess["user_id"] = uid
            sess["pilot_access"] = True

        csv_data = b"Date,From,To,Reg,SE,PIC\n15/03/24,EBBR,EBOS,OO-TST,0.8,0.8\n"
        rv1 = self._upload_csv(client, csv_data)
        assert rv1.status_code == 200

        rv2 = self._execute(client)
        assert rv2.status_code == 302

        with app.app_context():
            assert FlightEntry.query.filter_by(aircraft_id=other_ac_id).first() is None
            entry = PilotLogbookEntry.query.filter_by(pilot_user_id=uid).first()
            assert entry is not None
            assert entry.flight_id is None
